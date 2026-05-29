"""First-run setup wizard. Guides new users through initial configuration."""
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify
from flask_login import login_user, current_user
from app import db, csrf
from app.models.user import User
from app.models.student import Student
from app.utils.audit import log_action
import json, csv, io, os
from config import Config

setup_bp = Blueprint('setup', __name__, template_folder='../templates/setup')


def needs_setup():
    """Check if the app needs first-run setup."""
    user = User.query.first()
    return user is None or not user.setup_completed


@setup_bp.route('/setup', methods=['GET', 'POST'])
def index():
    """Multi-step setup wizard."""
    if not needs_setup() and current_user.is_authenticated:
        return redirect(url_for('dashboard.index'))

    if request.method == 'POST':
        step = request.form.get('step', '')
        if step == 'complete':
            return _handle_complete(request.form)

    return render_template('setup/wizard.html')


def _handle_complete(form):
    """Process the full setup form submission."""
    user = User.query.first()
    if not user:
        user = User(username='counselor', display_name='School Counselor', role='counselor')
        user.set_password('changeme')
        db.session.add(user)
        db.session.flush()

    # Profile
    display_name = form.get('display_name', '').strip()
    if display_name:
        user.display_name = display_name

    school_name = form.get('school_name', '').strip()
    if school_name:
        user.school_name = school_name

    username = form.get('username', '').strip()
    if username and len(username) >= 3:
        existing = User.query.filter_by(username=username).first()
        if not existing or existing.id == user.id:
            user.username = username

    password = form.get('password', '')
    if password and len(password) >= 8:
        user.set_password(password)

    # School config — merge all fields into JSON
    school_config = {}
    if user.school_config_json:
        try:
            school_config = json.loads(user.school_config_json)
        except (json.JSONDecodeError, TypeError):
            pass

    for key in ('counselor_title', 'school_year_start', 'school_year_end',
                'mascotEmoji', 'motto', 'shortName'):
        val = form.get(key, '').strip()
        if val:
            school_config[key] = val

    grade_levels = form.getlist('grade_levels')
    if grade_levels:
        school_config['grade_levels'] = grade_levels

    primary_color = form.get('primary_color', '').strip()
    secondary_color = form.get('secondary_color', '').strip()
    if primary_color:
        school_config.setdefault('colors', {})['primary'] = primary_color
    if secondary_color:
        school_config.setdefault('colors', {})['secondary'] = secondary_color

    contact_phone = form.get('contact_phone', '').strip()
    contact_email = form.get('contact_email', '').strip()
    contact_address = form.get('contact_address', '').strip()
    if contact_phone:
        school_config['contactPhone'] = contact_phone
    if contact_email:
        school_config['contactEmail'] = contact_email
    if contact_address:
        school_config['contactAddress'] = contact_address

    if school_name:
        school_config['schoolName'] = school_name

    # Ollama settings
    ollama_url = form.get('ollama_url', '').strip()
    ollama_model = form.get('ollama_model', '').strip()
    if ollama_url or ollama_model:
        try:
            from app.utils.ollama_client import save_settings
            save_settings(ollama_url or 'http://localhost:11434', ollama_model or 'gemma3:4b')
        except Exception:
            pass

    # iCal URL
    ical_url = form.get('ical_url', '').strip()
    if ical_url:
        user.external_ical_url = ical_url

    # Theme
    theme = form.get('theme', '').strip()
    if theme and theme in ('light', 'dark', 'school', 'focus', 'auto'):
        user.theme_preference = theme

    user.school_config_json = json.dumps(school_config)
    user.setup_completed = True
    db.session.commit()

    login_user(user, remember=False)
    log_action('setup_complete', 'user', user.id)

    flash('Setup complete! Welcome to Counselor Assistant.', 'success')
    return redirect(url_for('dashboard.index'))


@setup_bp.route('/setup/import-preview', methods=['POST'])
@csrf.exempt
def import_preview():
    """Preview CSV/Excel headers and sample data. Returns JSON."""
    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'ok': False, 'error': 'No file selected'}), 400

    fname = file.filename.lower()
    try:
        if fname.endswith('.csv'):
            content = file.read().decode('utf-8-sig')
            reader = csv.reader(io.StringIO(content))
            headers = next(reader)
            rows = []
            for i, row in enumerate(reader):
                if i < 5:
                    rows.append(row)
            row_count = i + 1 if 'i' in dir() else 0
            # Re-read to count all
            content_reader = csv.reader(io.StringIO(content))
            next(content_reader)
            row_count = sum(1 for _ in content_reader)
        elif fname.endswith(('.xlsx', '.xls')):
            from openpyxl import load_workbook
            wb = load_workbook(file, data_only=True, read_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                return jsonify({'ok': False, 'error': 'Empty file'}), 400
            headers = [str(c) if c else '' for c in all_rows[0]]
            rows = [[str(c) if c else '' for c in r] for r in all_rows[1:6]]
            row_count = len(all_rows) - 1
            wb.close()
        else:
            return jsonify({'ok': False, 'error': 'Please upload a CSV or Excel file'}), 400

        headers = [h.strip() for h in headers if h.strip()]
        return jsonify({'ok': True, 'headers': headers, 'row_count': row_count, 'sample': rows})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400


@setup_bp.route('/setup/import-students', methods=['POST'])
@csrf.exempt
def import_students():
    """Import students from CSV/Excel during setup. Returns JSON."""
    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'ok': False, 'error': 'No file selected'}), 400

    mapping_json = request.form.get('mapping', '{}')
    try:
        mapping = json.loads(mapping_json)
    except (json.JSONDecodeError, TypeError):
        return jsonify({'ok': False, 'error': 'Invalid column mapping'}), 400

    user = User.query.first()
    if not user:
        return jsonify({'ok': False, 'error': 'No user found'}), 400

    fname = file.filename.lower()
    try:
        if fname.endswith('.csv'):
            content = file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
            rows = list(reader)
        elif fname.endswith(('.xlsx', '.xls')):
            from openpyxl import load_workbook
            wb = load_workbook(file, data_only=True, read_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if len(all_rows) < 2:
                return jsonify({'ok': False, 'error': 'No data rows found'}), 400
            headers = [str(c).strip() if c else '' for c in all_rows[0]]
            rows = []
            for r in all_rows[1:]:
                row_dict = {}
                for j, h in enumerate(headers):
                    if h and j < len(r):
                        row_dict[h] = str(r[j]) if r[j] is not None else ''
                rows.append(row_dict)
            wb.close()
        else:
            return jsonify({'ok': False, 'error': 'Unsupported file type'}), 400
    except Exception as e:
        return jsonify({'ok': False, 'error': f'File read error: {e}'}), 400

    imported = 0
    skipped = 0
    for row in rows:
        first_name = row.get(mapping.get('first_name', ''), '').strip()
        last_name = row.get(mapping.get('last_name', ''), '').strip()
        student_id = row.get(mapping.get('student_id', ''), '').strip()
        grade_str = row.get(mapping.get('grade_level', ''), '').strip()

        if not first_name and not last_name:
            skipped += 1
            continue

        if not student_id:
            student_id = f'SETUP-{imported + skipped + 1}'

        existing = Student.query.filter_by(
            student_id_number=student_id,
            assigned_counselor_id=user.id
        ).first()
        if existing:
            skipped += 1
            continue

        grade_level = None
        if grade_str:
            try:
                grade_level = int(grade_str.replace('th', '').replace('st', '').replace('nd', '').replace('rd', '').strip())
            except ValueError:
                pass

        student = Student(
            first_name=first_name or 'Unknown',
            last_name=last_name or 'Unknown',
            student_id_number=student_id,
            grade_level=grade_level,
            assigned_counselor_id=user.id,
            status='active',
        )

        # Optional fields from mapping
        for field_key, col_name in mapping.items():
            if field_key in ('first_name', 'last_name', 'student_id', 'grade_level'):
                continue
            val = row.get(col_name, '').strip()
            if not val:
                continue
            if field_key == 'gender' and hasattr(student, 'gender'):
                student.gender = val
            elif field_key == 'email' and hasattr(student, 'email'):
                student.email = val
            elif field_key == 'ethnicity' and hasattr(student, 'ethnicity'):
                student.ethnicity = val
            elif field_key == 'date_of_birth' and hasattr(student, 'date_of_birth'):
                from datetime import datetime
                for fmt in ('%m/%d/%Y', '%Y-%m-%d', '%m-%d-%Y'):
                    try:
                        student.date_of_birth = datetime.strptime(val, fmt).date()
                        break
                    except ValueError:
                        continue

        db.session.add(student)
        imported += 1

    db.session.commit()
    return jsonify({'ok': True, 'imported': imported, 'skipped': skipped})


@setup_bp.route('/setup/upload-logo', methods=['POST'])
@csrf.exempt
def upload_logo():
    """Upload school logo during setup. Returns JSON."""
    file = request.files.get('logo')
    if not file or not file.filename:
        return jsonify({'ok': False, 'error': 'No file selected'}), 400

    allowed = {'png', 'jpg', 'jpeg', 'svg', 'webp'}
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in allowed:
        return jsonify({'ok': False, 'error': 'Allowed: PNG, JPG, SVG, WebP'}), 400

    user = User.query.first()
    uid = user.id if user else 1

    logo_dir = os.path.join(Config.BASE_DIR, 'data', 'uploads', 'school_logos')
    os.makedirs(logo_dir, exist_ok=True)

    # Remove any existing logo
    import glob
    for old in glob.glob(os.path.join(logo_dir, f'logo_{uid}.*')):
        os.remove(old)

    filename = f'logo_{uid}.{ext}'
    filepath = os.path.join(logo_dir, filename)
    file.save(filepath)

    logo_url = f'/course-catalog/api/school-logo'
    return jsonify({'ok': True, 'logoUrl': logo_url})
