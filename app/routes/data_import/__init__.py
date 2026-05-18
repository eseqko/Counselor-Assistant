"""Routes for importing attendance, grades, and other student data via CSV/Excel.

Split into sub-modules by domain:
    index        — Import hub page
    attendance   — Attendance template/upload/clear
    grades       — Grades template/preview/upload/clear
    students     — Student info bulk update template/upload
    _parsers     — Shared file-parsing helpers (CSV/Excel, Synergy format)

The blueprint and cross-cutting constants (openpyxl shim, validation sets,
Synergy code map, updatable student fields) live in this file. Sub-modules
import them via `from app.routes.data_import import data_import_bp, ...`.
"""
from flask import Blueprint

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    HAS_OPENPYXL = True
except ImportError:
    Workbook = load_workbook = None
    Font = PatternFill = Alignment = Border = Side = None
    get_column_letter = None
    DataValidation = None
    HAS_OPENPYXL = False

data_import_bp = Blueprint('data_import', __name__)

VALID_ATTENDANCE = {'present', 'absent', 'tardy', 'excused'}
VALID_GRADES = {'A+', 'A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-',
                'D+', 'D', 'D-', 'F', 'P', 'NP', 'I', 'W'}

# Synergy SIS attendance code → (status, reason)
SYNERGY_STATUS_MAP = {
    '':                 ('present', ''),
    'activity':         ('excused', 'Activity'),
    'illness':          ('excused', 'Illness'),
    'excused':          ('excused', 'Excused'),
    'counseling':       ('excused', 'Counseling'),
    'testing':          ('excused', 'Testing'),
    'office excused':   ('excused', 'Office Excused'),
    'office ex':        ('excused', 'Office Excused'),
    'cut':              ('absent', 'Cut'),
    'unverified':       ('absent', 'Unverified'),
    'parent unexcused': ('absent', 'Parent Unexcused'),
    'tardy':            ('tardy', 'Tardy'),
    'unexcused tardy':  ('tardy', 'Unexcused Tardy'),
}

# Fields that can be bulk-updated via spreadsheet
STUDENT_UPDATE_FIELDS = {
    'first_name':              ('First Name', str),
    'last_name':                ('Last Name', str),
    'grade_level':              ('Grade Level', int),
    'gender':                   ('Gender', str),
    'ethnicity':                ('Ethnicity', str),
    'email':                    ('Student Email', str),
    'phone':                    ('Phone', str),
    'date_of_birth':            ('Date of Birth', 'date'),
    'parent_guardian_name':     ('Parent/Guardian', str),
    'parent_guardian_phone':    ('Parent Phone', str),
    'parent_guardian_email':    ('Parent Email', str),
    'address':                  ('Address', str),
    'homeroom':                 ('Advisory', str),
    'el_status':                ('EL Status', str),
    'el_level':                 ('EL Level', str),
    'us_school_entry_date':     ('US School Entry Date', 'date'),
    'iep_status':               ('IEP', 'bool'),
    'section_504':              ('504', 'bool'),
}

# Import sub-modules at the bottom so they can register routes on the blueprint
from app.routes.data_import import index, attendance, grades, students, elpac  # noqa: E402, F401
