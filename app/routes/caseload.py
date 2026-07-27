import io
import json
from datetime import date, datetime, timezone
from flask import (Blueprint, render_template, request, redirect, url_for,
                   flash, send_file, jsonify, abort)
from flask_login import login_required, current_user
from sqlalchemy.exc import IntegrityError
from app import db, csrf
from app.models.student import Student, Tag
from app.models.rollover import RolloverSnapshot
from app.models.transcript import TranscriptRecord
from app.utils.audit import log_action
from app.utils.helpers import parse_date
from app.utils.roles import owned_or_404


def _safe_float(value, default=0.0):
    """Coerce a possibly-stringy/None value to float; fall back on garbage.

    Transcript-batch data comes from a client-side PDF parser, so numeric fields
    can arrive as strings or junk tokens. SQLite would store them verbatim in a
    Float column and crash later during credit math — coerce at the boundary.
    """
    try:
        return float(value)
    except (TypeError, ValueError):
        return default
from app.routes.graduation import (
    STATE_MIN_REQUIREMENTS, STATE_MIN_TOTAL, TOTAL_REQUIRED, _risk_level,
)
from app.utils.cte_status import compute_cte_status
from app.utils import rollover as rollover_util
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

caseload_bp = Blueprint('caseload', __name__)

# ---------- EL status constants ----------
VALID_EL_STATUSES = {'Newcomer', 'LTEL', 'RFEP', 'EO', ''}
VALID_EL_LEVELS = {'EL 1', 'EL 2', 'EL 3', ''}


@caseload_bp.route('/')
@login_required
def index():
    search = request.args.get('search', '').strip()
    grade = request.args.get('grade', '')
    status = request.args.get('status', 'active')
    tag_filter = request.args.get('tag', '')
    el_filter = request.args.get('el_status', '')

    # Exclude the per-user Sample Student (screener test vehicle) from the roster.
    query = Student.query.filter_by(assigned_counselor_id=current_user.id).filter(
        Student.is_sample == False)

    if status:
        query = query.filter_by(status=status)
    if grade:
        query = query.filter_by(grade_level=int(grade))
    if search:
        query = query.filter(
            db.or_(
                Student.first_name.ilike(f'%{search}%'),
                Student.last_name.ilike(f'%{search}%'),
                Student.student_id_number.ilike(f'%{search}%'),
            )
        )
    if tag_filter:
        query = query.filter(Student.tags.any(Tag.name == tag_filter))
    if el_filter:
        query = query.filter_by(el_status=el_filter)

    page = request.args.get('page', 1, type=int)
    pagination = query.order_by(
        Student.last_name, Student.first_name
    ).paginate(page=max(1, page), per_page=50, error_out=False)
    tags = Tag.query.order_by(Tag.name).all()

    return render_template('caseload/index.html',
        students=pagination.items, pagination=pagination,
        search=search, grade=grade,
        status=status, tag_filter=tag_filter, el_filter=el_filter, tags=tags)


@caseload_bp.route('/add', methods=['GET', 'POST'])
@login_required
def add_student():
    if request.method == 'POST':
        el_status = request.form.get('el_status', 'EO')
        el_level = request.form.get('el_level', '') if el_status == 'Newcomer' else ''

        student = Student(
            student_id_number=request.form['student_id_number'],
            first_name=request.form['first_name'],
            last_name=request.form['last_name'],
            grade_level=int(request.form['grade_level']) if request.form.get('grade_level') else None,
            date_of_birth=parse_date(request.form.get('date_of_birth')),
            gender=request.form.get('gender', ''),
            ethnicity=request.form.get('ethnicity', ''),
            email=request.form.get('email', ''),
            phone=request.form.get('phone', ''),
            parent_guardian_name=request.form.get('parent_guardian_name', ''),
            parent_guardian_phone=request.form.get('parent_guardian_phone', ''),
            parent_guardian_email=request.form.get('parent_guardian_email', ''),
            address=request.form.get('address', ''),
            homeroom=request.form.get('homeroom', ''),
            assigned_counselor_id=current_user.id,
            iep_status='iep_status' in request.form,
            section_504='section_504' in request.form,
            el_status=el_status,
            el_level=el_level,
            ell_status=(el_status in ('Newcomer', 'LTEL', 'RFEP')),
            enrollment_date=parse_date(request.form.get('enrollment_date')),
            is_foster_youth='is_foster_youth' in request.form,
            is_homeless='is_homeless' in request.form,
            is_migrant_newcomer='is_migrant_newcomer' in request.form,
            is_formerly_incarcerated='is_formerly_incarcerated' in request.form,
            is_military_connected='is_military_connected' in request.form,
            ab_exemption_status=request.form.get('ab_exemption_status', 'none'),
            ab_transfer_date=parse_date(request.form.get('ab_transfer_date')),
            ab_exemption_date=parse_date(request.form.get('ab_exemption_date')),
        )
        # Handle tags — bulk-load existing names in one query
        wanted = {n.strip() for n in request.form.get('tags', '').split(',') if n.strip()}
        if wanted:
            existing = {t.name: t for t in Tag.query.filter(Tag.name.in_(wanted)).all()}
            for name in wanted:
                tag = existing.get(name)
                if not tag:
                    tag = Tag(name=name)
                    db.session.add(tag)
                student.tags.append(tag)

        db.session.add(student)
        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash(f"A student with ID '{student.student_id_number}' already exists. "
                  "Please use a unique student ID.", 'danger')
            tags = Tag.query.order_by(Tag.name).all()
            return render_template('caseload/add.html', tags=tags,
                el_statuses=Student.EL_STATUSES, el_levels=Student.EL_LEVELS,
                ab_populations=Student.AB_POPULATION_FIELDS,
                ab_statuses=Student.AB_EXEMPTION_STATUSES)
        log_action('create', 'student', student.id, f'Added student {student.full_name}')
        flash(f'Student {student.full_name} added successfully.', 'success')
        return redirect(url_for('caseload.view_student', id=student.id))

    tags = Tag.query.order_by(Tag.name).all()
    return render_template('caseload/add.html', tags=tags,
        el_statuses=Student.EL_STATUSES, el_levels=Student.EL_LEVELS,
        ab_populations=Student.AB_POPULATION_FIELDS,
        ab_statuses=Student.AB_EXEMPTION_STATUSES)


@caseload_bp.route('/<int:id>')
@login_required
def view_student(id):
    student = owned_or_404(Student, id, 'assigned_counselor_id')
    log_action('view', 'student', student.id)
    notes = student.notes.limit(10).all()
    latest_transcript = student.transcript_records.first()

    uses_state_min = student.uses_state_minimum
    total_required = STATE_MIN_TOTAL if uses_state_min else TOTAL_REQUIRED

    # Pre-parse JSON fields for template
    transcript_credits = None
    transcript_ag = None
    credits_total_shortfall = 0
    credits_all_met = True
    state_min_risk = None
    if latest_transcript:
        if latest_transcript.credits_json:
            try:
                transcript_credits = json.loads(latest_transcript.credits_json)
                if uses_state_min:
                    # AB exemption accepted: filter to CA state minimum subjects only
                    # and override per-subject required amounts (Ed Code 51225.3).
                    adapted = {}
                    for subj, req in STATE_MIN_REQUIREMENTS.items():
                        data = transcript_credits.get(subj, {}) or {}
                        comp = data.get('completed', 0) or 0
                        wip = data.get('wip', 0) or 0
                        adapted[subj] = {
                            'required': req,
                            'completed': comp,
                            'wip': wip,
                            'need': max(0, req - comp),
                        }
                    transcript_credits = adapted
                    # Recompute risk vs state min (a student "critical" at 225 may be
                    # "on-track" at 130 with the same credits earned).
                    state_min_risk = _risk_level(
                        latest_transcript.total_completed or 0,
                        STATE_MIN_TOTAL, student.grade_level)
                # Sum per-subject shortfalls using whichever set is active
                for data in transcript_credits.values():
                    req = data.get('required', 0) or 0
                    comp = data.get('completed', 0) or 0
                    if comp < req:
                        credits_all_met = False
                        credits_total_shortfall += req - comp
            except (json.JSONDecodeError, TypeError):
                pass
        if latest_transcript.ag_json:
            try:
                transcript_ag = json.loads(latest_transcript.ag_json)
            except (json.JSONDecodeError, TypeError):
                pass

    cte_details = None
    if latest_transcript and latest_transcript.cte_courses_json:
        try:
            cte_details = json.loads(latest_transcript.cte_courses_json)
        except (json.JSONDecodeError, TypeError):
            pass

    cte_status = compute_cte_status(
        cte_details,
        latest_transcript.cte_completed if latest_transcript else 0,
    )

    goals_completion_pct = None
    total_goals = student.goals.count()
    if total_goals:
        achieved = student.goals.filter_by(status='achieved').count()
        goals_completion_pct = round(100 * achieved / total_goals, 1)

    # ── Student 360: action plan + trends ────────────────────────────────
    # The plan consolidates on graduation's engine (the single home for
    # credit math) and computes the on-track verdict, prioritized next steps,
    # attendance with the CORRECTED day-based denominator (the old
    # present-records / total-period-records math understated absence), GPA
    # trajectory, and credit velocity across transcript imports.
    from app.routes.graduation import _build_student_grad_data
    from app.utils.next_steps import build_action_plan
    grad_data = _build_student_grad_data(student)
    action_plan = build_action_plan(student, grad_data=grad_data)
    att = action_plan['attendance']
    attendance_rate = (round(100 - att['rate_pct'], 1)
                       if att['rate_pct'] is not None else None)

    # Per-student screening history with delta vs the previous result of the
    # same screener (screenings previously had no per-student surface at all).
    from app.models.screening import ScreeningResult
    screening_rows = ScreeningResult.query.filter_by(
        student_id=student.id).order_by(
        ScreeningResult.administered_date.desc()).limit(12).all()
    prev_by_template = {}
    screenings_with_delta = []
    for r in reversed(screening_rows):          # oldest→newest to compute deltas
        prev = prev_by_template.get(r.template_id)
        delta = (r.total_score - prev.total_score
                 if prev and r.total_score is not None and prev.total_score is not None
                 else None)
        screenings_with_delta.append({'r': r, 'delta': delta})
        prev_by_template[r.template_id] = r
    screenings_with_delta.reverse()             # newest first for display

    return render_template('caseload/view.html',
        student=student, notes=notes,
        latest_transcript=latest_transcript,
        transcript_credits=transcript_credits,
        transcript_ag=transcript_ag,
        cte_details=cte_details,
        cte_status=cte_status,
        credits_total_shortfall=credits_total_shortfall,
        credits_all_met=credits_all_met,
        total_required=total_required,
        uses_state_min=uses_state_min,
        state_min_risk=state_min_risk,
        attendance_rate=attendance_rate,
        goals_completion_pct=goals_completion_pct,
        action_plan=action_plan,
        grad_data=grad_data,
        screenings_with_delta=screenings_with_delta,
        now_date=date.today(),
        exit_reasons=Student.EXIT_REASONS)


@caseload_bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_student(id):
    student = owned_or_404(Student, id, 'assigned_counselor_id')

    if request.method == 'POST':
        student.student_id_number = request.form['student_id_number']
        student.first_name = request.form['first_name']
        student.last_name = request.form['last_name']
        student.grade_level = int(request.form['grade_level']) if request.form.get('grade_level') else None
        student.date_of_birth = parse_date(request.form.get('date_of_birth'))
        student.gender = request.form.get('gender', '')
        student.ethnicity = request.form.get('ethnicity', '')
        student.email = request.form.get('email', '')
        student.phone = request.form.get('phone', '')
        student.parent_guardian_name = request.form.get('parent_guardian_name', '')
        student.parent_guardian_phone = request.form.get('parent_guardian_phone', '')
        student.parent_guardian_email = request.form.get('parent_guardian_email', '')
        student.address = request.form.get('address', '')
        student.homeroom = request.form.get('homeroom', '')
        student.status = request.form.get('status', 'active')
        student.iep_status = 'iep_status' in request.form
        student.section_504 = 'section_504' in request.form
        student.el_status = request.form.get('el_status', 'EO')
        student.el_level = request.form.get('el_level', '') if student.el_status == 'Newcomer' else ''
        student.ell_status = (student.el_status in ('Newcomer', 'LTEL', 'RFEP'))

        # AB Graduation Exemption fields
        student.is_foster_youth = 'is_foster_youth' in request.form
        student.is_homeless = 'is_homeless' in request.form
        student.is_migrant_newcomer = 'is_migrant_newcomer' in request.form
        student.is_formerly_incarcerated = 'is_formerly_incarcerated' in request.form
        student.is_military_connected = 'is_military_connected' in request.form
        student.ab_exemption_status = request.form.get('ab_exemption_status', 'none')
        student.ab_transfer_date = parse_date(request.form.get('ab_transfer_date'))
        student.ab_exemption_date = parse_date(request.form.get('ab_exemption_date'))

        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash(f"A student with ID '{student.student_id_number}' already exists. "
                  "Please use a unique student ID.", 'danger')
            tags = Tag.query.order_by(Tag.name).all()
            return render_template('caseload/edit.html', student=student, tags=tags,
                el_statuses=Student.EL_STATUSES, el_levels=Student.EL_LEVELS,
                ab_populations=Student.AB_POPULATION_FIELDS,
                ab_statuses=Student.AB_EXEMPTION_STATUSES)
        log_action('update', 'student', student.id, f'Updated student {student.full_name}')
        flash(f'Student {student.full_name} updated.', 'success')
        return redirect(url_for('caseload.view_student', id=student.id))

    tags = Tag.query.order_by(Tag.name).all()
    return render_template('caseload/edit.html', student=student, tags=tags,
        el_statuses=Student.EL_STATUSES, el_levels=Student.EL_LEVELS,
        ab_populations=Student.AB_POPULATION_FIELDS,
        ab_statuses=Student.AB_EXEMPTION_STATUSES)


@caseload_bp.route('/<int:id>/delete', methods=['POST'])
@login_required
def delete_student(id):
    student = owned_or_404(Student, id, 'assigned_counselor_id')
    exit_reason = request.form.get('exit_reason', 'other')
    exit_notes = request.form.get('exit_notes', '').strip()

    # Map exit reason to a status
    status_map = {
        'graduated': 'graduated',
        'transferred_in_district': 'transferred',
        'transferred_out_district': 'transferred',
        'counselor_change': 'transferred',
        'dropped_out': 'inactive',
        'aged_out': 'inactive',
        'expelled': 'inactive',
        'other': 'inactive',
    }
    student.status = status_map.get(exit_reason, 'inactive')
    student.exit_reason = exit_reason
    student.exit_date = date.today()
    student.exit_notes = exit_notes or None

    reason_label = dict(Student.EXIT_REASONS).get(exit_reason, exit_reason)
    log_action('remove', 'student', student.id,
               f'Removed student {student.full_name} — Reason: {reason_label}')
    db.session.commit()
    flash(f'{student.full_name} removed from caseload ({reason_label}).', 'warning')
    return redirect(url_for('caseload.index'))


# =====================================================================
#  EXCEL TEMPLATE DOWNLOAD
# =====================================================================

@caseload_bp.route('/download-template')
@login_required
def download_template():
    """Generate and download a formatted Excel template for caseload upload."""
    if not HAS_OPENPYXL:
        flash('Excel support requires the openpyxl package. Install it with: pip install openpyxl', 'danger')
        return redirect(url_for('caseload.index'))
    wb = Workbook()
    ws = wb.active
    ws.title = "Caseload Import"

    # --- Styles ---
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='2C5F8A', end_color='2C5F8A', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    locked = Protection(locked=True)
    unlocked = Protection(locked=False)

    # --- Column definitions ---
    columns = [
        ('First Name', 20, 'Enter student first name'),
        ('Last Name', 20, 'Enter student last name'),
        ('Grade', 8, 'Grade level (6-12)'),
        ('Student ID #', 16, 'Unique school student ID'),
        ('Email', 30, 'Student email address'),
        ('EL Status', 16, 'Newcomer, LTEL, RFEP, or EO'),
        ('EL Level', 12, 'Only if Newcomer: EL 1, EL 2, or EL 3'),
        ('IEP', 8, 'Yes or leave blank'),
        ('504 Plan', 10, 'Yes or leave blank'),
    ]

    # --- Write headers ---
    for col_idx, (name, width, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        cell.protection = locked
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Data validation ---
    # Grade: 6-12
    grade_dv = DataValidation(
        type='list', formula1='"6,7,8,9,10,11,12"', allow_blank=True,
        showErrorMessage=True, errorTitle='Invalid Grade',
        error='Please enter a grade from 6-12.'
    )
    grade_dv.sqref = 'C2:C1000'
    ws.add_data_validation(grade_dv)

    # EL Status: Newcomer, LTEL, RFEP, EO
    el_dv = DataValidation(
        type='list', formula1='"Newcomer,LTEL,RFEP,EO"', allow_blank=True,
        showErrorMessage=True, errorTitle='Invalid EL Status',
        error='Please choose: Newcomer, LTEL, RFEP, or EO'
    )
    el_dv.sqref = 'F2:F1000'
    ws.add_data_validation(el_dv)

    # EL Level: EL 1, EL 2, EL 3 (only for Newcomers)
    level_dv = DataValidation(
        type='list', formula1='"EL 1,EL 2,EL 3"', allow_blank=True,
        showErrorMessage=True, errorTitle='Invalid EL Level',
        error='Please choose: EL 1, EL 2, or EL 3 (only for Newcomer students)'
    )
    level_dv.sqref = 'G2:G1000'
    ws.add_data_validation(level_dv)

    # IEP: Yes or blank
    iep_dv = DataValidation(
        type='list', formula1='"Yes"', allow_blank=True,
        showErrorMessage=True, errorTitle='Invalid IEP',
        error='Enter "Yes" or leave blank.'
    )
    iep_dv.sqref = 'H2:H1000'
    ws.add_data_validation(iep_dv)

    # 504: Yes or blank
    plan_dv = DataValidation(
        type='list', formula1='"Yes"', allow_blank=True,
        showErrorMessage=True, errorTitle='Invalid 504 Plan',
        error='Enter "Yes" or leave blank.'
    )
    plan_dv.sqref = 'I2:I1000'
    ws.add_data_validation(plan_dv)

    # --- Format data rows (light alternating) ---
    alt_fill = PatternFill(start_color='F0F6FF', end_color='F0F6FF', fill_type='solid')
    for row in range(2, 52):  # Pre-format 50 rows
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = thin_border
            cell.protection = unlocked
            cell.alignment = Alignment(vertical='center')
            if row % 2 == 0:
                cell.fill = alt_fill

    # --- Instructions sheet ---
    instr = wb.create_sheet('Instructions')
    instr.sheet_properties.tabColor = 'E8A838'

    instructions = [
        ('CASELOAD UPLOAD TEMPLATE - INSTRUCTIONS', ''),
        ('', ''),
        ('Column', 'Instructions'),
        ('First Name', 'Required. Student\'s first name.'),
        ('Last Name', 'Required. Student\'s last name.'),
        ('Grade', 'Required. Grade level from 6-12. Use the dropdown.'),
        ('Student ID #', 'Required. Must be unique. This is the school\'s student ID number.'),
        ('Email', 'Optional. Student email address.'),
        ('EL Status', 'Required. Choose from dropdown: Newcomer, LTEL, RFEP, or EO (English Only).'),
        ('EL Level', 'Only fill in if EL Status is "Newcomer". Choose: EL 1, EL 2, or EL 3.'),
        ('IEP', 'Enter "Yes" if student has an IEP. Otherwise leave blank.'),
        ('504 Plan', 'Enter "Yes" if student has a 504 Plan. Otherwise leave blank.'),
        ('', ''),
        ('NOTES:', ''),
        ('', 'Duplicate Student IDs will update the existing student record (not create duplicates).'),
        ('', 'EL Level is ONLY for Newcomer students. It will be ignored for other EL statuses.'),
        ('', 'All data stays local on your computer. Nothing is uploaded to the cloud.'),
    ]

    title_font = Font(name='Calibri', bold=True, size=14, color='2C5F8A')
    bold_font = Font(name='Calibri', bold=True, size=11)
    normal_font = Font(name='Calibri', size=11)

    for row_idx, (col_a, col_b) in enumerate(instructions, 1):
        cell_a = instr.cell(row=row_idx, column=1, value=col_a)
        cell_b = instr.cell(row=row_idx, column=2, value=col_b)
        if row_idx == 1:
            cell_a.font = title_font
        elif row_idx == 3:
            cell_a.font = bold_font
            cell_b.font = bold_font
        else:
            cell_a.font = bold_font if col_a else normal_font
            cell_b.font = normal_font

    instr.column_dimensions['A'].width = 18
    instr.column_dimensions['B'].width = 75

    # --- Freeze top row and protect ---
    ws.freeze_panes = 'A2'
    ws.protection.sheet = True
    ws.protection.formatCells = False
    ws.protection.insertRows = False
    ws.protection.sort = False
    ws.protection.autoFilter = False

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    log_action('export', 'caseload_template', details='Downloaded caseload template')

    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='Caseload_Upload_Template.xlsx'
    )


# =====================================================================
#  EXCEL UPLOAD
# =====================================================================

def _parse_caseload_file(file):
    """Parse + validate a caseload template upload.

    Returns (rows, errors, fatal): `rows` are normalized dicts, `errors` are
    per-row messages, `fatal` is a whole-file error string (bad file / wrong
    headers) or None. Shared by the preview and apply routes so the diff the
    counselor previews is computed by the exact code that later applies it.
    """
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        return [], [], 'Please upload an Excel file (.xlsx).'
    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active
    except Exception as e:
        return [], [], f'Could not read Excel file: {str(e)}'

    expected = ['first name', 'last name', 'grade', 'student id #', 'email',
                'el status', 'el level', 'iep', '504 plan']
    headers = [str(cell.value or '').strip().lower() for cell in ws[1]]
    if headers[:len(expected)] != expected:
        return [], [], ('Column headers don\'t match the template. '
                        'Please download a fresh template and try again.')

    rows, errors = [], []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=9, values_only=True), start=2):
        first_name, last_name, grade, student_id, email, el_status, el_level, iep, plan_504 = row

        # Skip empty rows
        if not first_name and not last_name and not student_id:
            continue

        row_errors = []
        if not first_name:
            row_errors.append('First Name is required')
        if not last_name:
            row_errors.append('Last Name is required')
        if not student_id:
            row_errors.append('Student ID # is required')
        if not grade:
            row_errors.append('Grade is required')

        grade_val = None
        if grade:
            try:
                grade_val = int(grade)
                if grade_val < 6 or grade_val > 12:
                    row_errors.append(f'Grade must be 6-12, got {grade_val}')
            except (ValueError, TypeError):
                row_errors.append(f'Invalid grade: {grade}')

        el_status_clean = str(el_status or '').strip()
        if el_status_clean and el_status_clean not in VALID_EL_STATUSES:
            row_errors.append(f'Invalid EL Status: {el_status_clean}. Must be Newcomer, LTEL, RFEP, or EO.')
        if not el_status_clean:
            el_status_clean = 'EO'

        el_level_clean = str(el_level or '').strip()
        if el_status_clean == 'Newcomer' and el_level_clean and el_level_clean not in VALID_EL_LEVELS:
            row_errors.append(f'Invalid EL Level: {el_level_clean}. Must be EL 1, EL 2, or EL 3.')
        if el_status_clean != 'Newcomer':
            el_level_clean = ''

        if row_errors:
            errors.append(f'Row {row_idx}: ' + '; '.join(row_errors))
            continue

        rows.append({
            'row_idx': row_idx,
            'sid': str(student_id).strip(),
            'first_name': str(first_name).strip(),
            'last_name': str(last_name).strip(),
            'grade_level': grade_val,
            'email': str(email or '').strip(),
            'el_status': el_status_clean,
            'el_level': el_level_clean,
            'iep': str(iep or '').strip().lower() in ('yes', 'y', 'true', '1'),
            'plan_504': str(plan_504 or '').strip().lower() in ('yes', 'y', 'true', '1'),
        })
    return rows, errors, None


@caseload_bp.route('/upload/preview', methods=['POST'])
@csrf.exempt
@login_required
def upload_caseload_preview():
    """Diff an uploaded roster against the current caseload WITHOUT applying.

    Returns the three-bucket JSON the new-year sync UI renders:
      returning — in file AND currently on my caseload (with grade changes)
      new       — in file, not on my caseload (brand-new / promotable / other-counselor)
      departing — on my active caseload but absent from the file
    """
    rows, errors, fatal = _parse_caseload_file(request.files.get('file'))
    if fatal:
        return jsonify({'ok': False, 'error': fatal}), 400

    # Dedupe file rows by student ID (last row wins, matching apply order).
    by_sid = {}
    for r in rows:
        by_sid[r['sid']] = r

    # My current active caseload (samples excluded — they're test fixtures).
    mine = {s.student_id_number: s for s in Student.query.filter_by(
        assigned_counselor_id=current_user.id, status='active',
        is_sample=False).all()}

    # Global lookup for the file's IDs (classify new vs promotable vs blocked).
    existing_global = {}
    if by_sid:
        sid_list = list(by_sid.keys())
        for chunk_start in range(0, len(sid_list), 900):
            chunk = sid_list[chunk_start:chunk_start + 900]
            for s in Student.query.filter(Student.student_id_number.in_(chunk)).all():
                existing_global[s.student_id_number] = s

    returning, new_students = [], []
    for sid, r in by_sid.items():
        cur = mine.get(sid)
        if cur:
            returning.append({
                'sid': sid, 'name': f'{r["last_name"]}, {r["first_name"]}',
                'grade_from': cur.grade_level, 'grade_to': r['grade_level'],
                'grade_changed': cur.grade_level != r['grade_level'],
            })
            continue
        g = existing_global.get(sid)
        if g is None:
            kind = 'brand_new'
        elif current_user.role != 'admin' \
                and g.assigned_counselor_id not in (None, 0, current_user.id):
            kind = 'other_counselor'   # will be skipped on apply (FERPA guard)
        else:
            kind = 'promotable'        # shadow / unassigned / re-activating
        new_students.append({
            'sid': sid, 'name': f'{r["last_name"]}, {r["first_name"]}',
            'grade': r['grade_level'], 'kind': kind,
        })

    departing = [
        {'id': s.id, 'sid': sid, 'name': f'{s.last_name}, {s.first_name}',
         'grade': s.grade_level}
        for sid, s in mine.items() if sid not in by_sid
    ]
    departing.sort(key=lambda d: d['name'].lower())
    returning.sort(key=lambda d: d['name'].lower())
    new_students.sort(key=lambda d: d['name'].lower())

    return jsonify({
        'ok': True,
        'returning': returning,
        'new': new_students,
        'departing': departing,
        'errors': errors,
        'counts': {
            'returning': len(returning),
            'new': len(new_students),
            'new_blocked': sum(1 for n in new_students if n['kind'] == 'other_counselor'),
            'departing': len(departing),
            'errors': len(errors),
        },
    })


@caseload_bp.route('/upload', methods=['GET', 'POST'])
@login_required
def upload_caseload():
    if not HAS_OPENPYXL:
        flash('Excel support requires the openpyxl package. Install it with: pip install openpyxl', 'danger')
        return redirect(url_for('caseload.index'))
    if request.method == 'POST':
        rows, errors, fatal = _parse_caseload_file(request.files.get('file'))
        if fatal:
            flash(fatal, 'danger')
            return redirect(url_for('caseload.upload_caseload'))

        added = 0
        updated = 0

        for r in rows:
            # student_id_number is globally unique (one row per student). Look it
            # up GLOBALLY: scoping the lookup to the current counselor would let a
            # colliding ID (owned by another counselor) fall through to an INSERT
            # that violates the unique constraint and aborts the entire batch.
            existing = Student.query.filter_by(student_id_number=r['sid']).first()

            if existing and current_user.role != 'admin' \
                    and existing.assigned_counselor_id not in (None, 0, current_user.id):
                # On another counselor's caseload: don't reassign via import (FERPA)
                # and don't crash the batch. Skip with a clear, per-row message.
                errors.append(
                    f'Row {r["row_idx"]}: student ID {r["sid"]} is on another '
                    "counselor's caseload — skipped (use Reassign to move).")
                continue

            if existing:
                existing.first_name = r['first_name']
                existing.last_name = r['last_name']
                existing.grade_level = r['grade_level']
                existing.email = r['email']
                existing.el_status = r['el_status']
                existing.el_level = r['el_level']
                existing.ell_status = (r['el_status'] in ('Newcomer', 'LTEL', 'RFEP'))
                existing.iep_status = r['iep']
                existing.section_504 = r['plan_504']
                existing.assigned_counselor_id = current_user.id
                # Promote any shadow record to a full caseload student so it stops
                # being filtered out of UI lists. Re-activate exited students who
                # reappear on a roster (e.g. re-enrolled after a transfer).
                existing.is_shadow = False
                if existing.status != 'active':
                    existing.status = 'active'
                    existing.exit_reason = None
                    existing.exit_date = None
                updated += 1
            else:
                student = Student(
                    student_id_number=r['sid'],
                    first_name=r['first_name'],
                    last_name=r['last_name'],
                    grade_level=r['grade_level'],
                    email=r['email'],
                    el_status=r['el_status'],
                    el_level=r['el_level'],
                    ell_status=(r['el_status'] in ('Newcomer', 'LTEL', 'RFEP')),
                    iep_status=r['iep'],
                    section_504=r['plan_504'],
                    assigned_counselor_id=current_user.id,
                    status='active',
                )
                db.session.add(student)
                added += 1

        # ── New-year sync: departing-student actions ──────────────────────
        # Optional JSON map {student_db_id: action} posted by the preview UI
        # for students on my caseload who are absent from the file. Absent map
        # or 'keep' = untouched (the safe default — a partial file can never
        # silently exit students). Applied atomically with the upserts, with a
        # RolloverSnapshot for the same 24-hour undo the EOY rollover has.
        departed_counts = {}
        snapshot = None
        raw_actions = request.form.get('departing_actions', '').strip()
        if raw_actions:
            try:
                action_map = {int(k): v for k, v in json.loads(raw_actions).items()}
            except (ValueError, TypeError):
                db.session.rollback()
                flash('Invalid departing-student actions payload — nothing was changed.', 'danger')
                return redirect(url_for('caseload.upload_caseload'))

            file_sids = {r['sid'] for r in rows}
            today = date.today()
            snapshot_items = []
            targets = Student.query.filter(
                Student.id.in_(list(action_map.keys())),
                Student.assigned_counselor_id == current_user.id,
                Student.is_sample == False,
            ).all() if action_map else []
            for s in targets:
                action = action_map.get(s.id)
                if action not in rollover_util.SYNC_ACTIONS or action == 'keep':
                    continue
                if s.student_id_number in file_sids:
                    # In the file after all (e.g. stale preview) — never exit a
                    # student the roster says is present.
                    continue
                prior = rollover_util.apply_sync_action(s, action, today)
                prior['applied_action'] = f'sync_{action}'
                snapshot_items.append(prior)
                departed_counts[action] = departed_counts.get(action, 0) + 1
                log_action('caseload_sync.apply', resource_type='student',
                           resource_id=s.id, details=f'action={action}')
            if snapshot_items:
                snapshot = RolloverSnapshot(
                    counselor_id=current_user.id,
                    student_count=len(snapshot_items),
                    school_year_end_date=today,
                    payload=json.dumps(snapshot_items),
                )
                db.session.add(snapshot)

        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash('Import aborted: a duplicate student ID conflict was detected. '
                  'No changes were saved.', 'danger')
            return render_template('caseload/upload.html',
                errors=['Duplicate student ID conflict — import aborted. No changes saved.'],
                added=0, updated=0)
        log_action('import', 'caseload', details=f'Imported caseload: {added} added, {updated} updated'
                   + (f', departing actions: {departed_counts}' if departed_counts else ''))

        msg = f'{added} students added, {updated} updated.'
        if departed_counts:
            pretty = ', '.join(f'{n} {a}' for a, n in sorted(departed_counts.items()))
            msg += f' Departing: {pretty}. Undo available for 24 hours on the Rollover page.'
        if errors:
            flash(f'Imported with issues: {msg} {len(errors)} errors.', 'warning')
            return render_template('caseload/upload.html', errors=errors, added=added, updated=updated)
        flash(f'Caseload imported successfully! {msg}', 'success')
        return redirect(url_for('caseload.index'))

    return render_template('caseload/upload.html', errors=None, added=0, updated=0)


# =====================================================================
#  EXPORT CURRENT CASELOAD TO EXCEL
# =====================================================================

@caseload_bp.route('/export')
@login_required
def export_caseload():
    """Export current caseload to Excel."""
    if not HAS_OPENPYXL:
        flash('Excel support requires the openpyxl package. Install it with: pip install openpyxl', 'danger')
        return redirect(url_for('caseload.index'))
    students = Student.query.filter_by(
        assigned_counselor_id=current_user.id
    ).order_by(Student.grade_level, Student.last_name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "My Caseload"

    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='2C5F8A', end_color='2C5F8A', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    headers = ['First Name', 'Last Name', 'Grade', 'Student ID #', 'Email',
               'EL Status', 'EL Level', 'IEP', '504 Plan', 'Status']
    widths = [18, 18, 8, 16, 30, 16, 12, 8, 10, 12]

    for col_idx, (name, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, s in enumerate(students, 2):
        values = [
            s.first_name, s.last_name, s.grade_level, s.student_id_number,
            s.email, s.el_status or 'EO', s.el_level or '',
            'Yes' if s.iep_status else '', 'Yes' if s.section_504 else '',
            s.status,
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    log_action('export', 'caseload', details=f'Exported {len(students)} students')

    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='My_Caseload_Export.xlsx'
    )


# =====================================================================
#  TRANSCRIPT ANALYSIS — SAVE & BATCH IMPORT
# =====================================================================

@caseload_bp.route('/transcript/batch')
@login_required
def transcript_batch():
    """Batch transcript import page."""
    return render_template('caseload/transcript_batch.html')


@caseload_bp.route('/transcript/save', methods=['POST'])
@login_required
def transcript_save():
    """Save transcript analysis results for one or more students.

    Expects JSON body: { "students": [ { permId, quarter, ... }, ... ] }
    Matches students by permId → student_id_number.
    """
    data = request.get_json()
    if not data or 'students' not in data:
        return jsonify({'error': 'Missing students data'}), 400

    saved = 0
    skipped = 0
    not_found = []

    for entry in data['students']:
        perm_id = str(entry.get('permId', '')).strip()
        if not perm_id:
            skipped += 1
            continue

        student = Student.query.filter_by(
            student_id_number=perm_id,
            assigned_counselor_id=current_user.id
        ).first()

        if not student:
            not_found.append(perm_id)
            continue

        record = TranscriptRecord(
            student_id=student.id,
            quarter=entry.get('quarter', ''),
            total_completed=_safe_float(entry.get('totalCompleted', 0)),
            total_wip=_safe_float(entry.get('totalWIP', 0)),
            total_needed=_safe_float(entry.get('totalNeeded', 0)),
            risk_level=entry.get('riskLevel', ''),
            ag_status=entry.get('agStatus', ''),
            ag_areas_met=int(_safe_float(entry.get('agAreasMet', 0))),
            ag_areas_deficient=int(_safe_float(entry.get('agAreasDeficient', 0))),
            cte_completed=_safe_float(entry.get('cteCompleted', 0)),
            cte_level=entry.get('cteLevel', 'none'),
            cte_is_completer=bool(entry.get('cteIsCompleter', False)),
            credits_json=json.dumps(entry.get('credits', {})),
            ag_json=json.dumps(entry.get('agAreas', {})),
            created_by_id=current_user.id,
        )
        db.session.add(record)
        saved += 1

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({'ok': False, 'error': 'Failed to save transcripts. No changes were made.'}), 500

    if saved > 0:
        log_action('import', 'transcript',
                   details=f'Saved transcript data for {saved} student(s)')

    return jsonify({
        'saved': saved,
        'skipped': skipped,
        'notFound': not_found,
        'message': f'Saved {saved} transcript record(s).'
                   + (f' {len(not_found)} student(s) not found in your caseload.' if not_found else '')
    })


@caseload_bp.route('/<int:id>/cte-courses', methods=['POST'])
@login_required
def save_cte_courses(id):
    student = Student.query.filter_by(
        id=id, assigned_counselor_id=current_user.id).first_or_404()
    latest = student.transcript_records.first()
    if not latest:
        flash('No transcript record found. Import a transcript first.', 'warning')
        return redirect(url_for('caseload.view_student', id=id))

    pathway = request.form.get('cte_pathway', '').strip()
    courses = []
    i = 0
    while True:
        name = request.form.get(f'cte_course_name_{i}')
        if name is None:
            break
        name = name.strip()
        if name:
            courses.append({
                'name': name,
                'grade': request.form.get(f'cte_course_grade_{i}', '').strip(),
                'credits': request.form.get(f'cte_course_credits_{i}', '10').strip(),
                'level': request.form.get(f'cte_course_level_{i}', '').strip(),
            })
        i += 1

    parsed = {'pathway': pathway, 'courses': courses}
    latest.cte_courses_json = json.dumps(parsed)
    # Recompute legacy fields so caseload-wide reports stay in sync.
    # cte_level shows the friendly district label; cte_is_completer follows
    # the strict Perkins V definition (what CALPADS would count).
    status = compute_cte_status(parsed, latest.cte_completed or 0)
    latest.cte_level = status['district_status']
    latest.cte_is_completer = (status['perkins_status'] == 'Completer')
    db.session.commit()
    flash('CTE pathway details saved.', 'success')
    return redirect(url_for('caseload.view_student', id=id))


# ── End-of-year rollover ─────────────────────────────────────────


@caseload_bp.route('/rollover')
@login_required
def rollover():
    """Review page: propose an action per student, editable inline."""
    students = (Student.query
                .filter_by(assigned_counselor_id=current_user.id, status='active')
                .order_by(Student.grade_level.desc().nulls_last(),
                          Student.last_name, Student.first_name)
                .all())

    rows = []
    anomalies = []
    for s in students:
        cs = rollover_util.credit_status_summary(s)
        action = rollover_util.default_action(s, credit_status=cs)
        flags = rollover_util.detect_anomalies(s, credit_status=cs)
        rows.append({
            'student': s,
            'default_action': action,
            'flags': flags,
        })
        if flags:
            anomalies.append({'student': s, 'flags': flags})

    # Default school-year-end: June 15 of the current calendar year (or next
    # year if we're already past June).
    today = date.today()
    end_year = today.year if today.month < 7 else today.year + 1
    default_end = date(end_year, 6, 15)

    # Recent snapshot still within undo window?
    recent_snapshot = (RolloverSnapshot.query
                       .filter_by(counselor_id=current_user.id, undone=False)
                       .order_by(RolloverSnapshot.created_at.desc())
                       .first())
    if recent_snapshot and recent_snapshot.is_expired():
        recent_snapshot = None

    return render_template('caseload/rollover.html',
                           rows=rows,
                           anomalies=anomalies,
                           actions=rollover_util.ACTIONS,
                           default_end_date=default_end,
                           recent_snapshot=recent_snapshot)


@caseload_bp.route('/rollover/confirm', methods=['POST'])
@login_required
def rollover_confirm():
    """Apply the per-row actions submitted from the review form."""
    end_date_str = request.form.get('end_date', '').strip()
    try:
        end_date = date.fromisoformat(end_date_str)
    except ValueError:
        flash('Invalid end-of-year date.', 'danger')
        return redirect(url_for('caseload.rollover'))

    # Map student_id -> chosen action. Anything missing or invalid is skipped.
    action_map = {}
    for key, value in request.form.items():
        if not key.startswith('action_'):
            continue
        try:
            sid = int(key[len('action_'):])
        except ValueError:
            continue
        if value in rollover_util.ACTION_KEYS:
            action_map[sid] = value

    if not action_map:
        flash('No actions submitted.', 'warning')
        return redirect(url_for('caseload.rollover'))

    students = (Student.query
                .filter(Student.id.in_(action_map.keys()),
                        Student.assigned_counselor_id == current_user.id)
                .all())

    counts = {}
    snapshot_items = []
    for s in students:
        action = action_map.get(s.id)
        if not action:
            continue
        prior = rollover_util.apply_action(s, action, end_date)
        prior['applied_action'] = action
        snapshot_items.append(prior)
        counts[action] = counts.get(action, 0) + 1
        log_action('rollover.apply', resource_type='student',
                   resource_id=s.id, details=f'action={action}')

    if not snapshot_items:
        flash('Nothing to apply — none of the selected students are on your caseload.', 'warning')
        return redirect(url_for('caseload.rollover'))

    snapshot = RolloverSnapshot(
        counselor_id=current_user.id,
        student_count=len(snapshot_items),
        school_year_end_date=end_date,
        payload=json.dumps(snapshot_items),
    )
    db.session.add(snapshot)
    db.session.commit()
    log_action('rollover.commit', resource_type='rollover_snapshot',
               resource_id=snapshot.id,
               details=f'students={len(snapshot_items)} end_date={end_date.isoformat()}')

    summary = ', '.join(f'{n} {a.replace("_", " ")}' for a, n in sorted(counts.items()))
    flash(f'Rollover applied: {summary}. Undo available for 24 hours.', 'success')
    return redirect(url_for('caseload.rollover'))


@caseload_bp.route('/rollover/undo/<int:snapshot_id>', methods=['POST'])
@login_required
def rollover_undo(snapshot_id):
    """Restore students to the state captured in the given snapshot."""
    snapshot = RolloverSnapshot.query.get_or_404(snapshot_id)
    if snapshot.counselor_id != current_user.id:
        abort(403)
    if not snapshot.can_undo():
        flash('That rollover can no longer be undone (already undone, or 24-hour window expired).', 'warning')
        return redirect(url_for('caseload.rollover'))

    restored = 0
    for entry in snapshot.items():
        s = Student.query.get(entry.get('student_id'))
        # Ownership: restore students who are mine NOW, or who are currently
        # UNOWNED but were mine at capture time (the sync 'unassign' action
        # clears assigned_counselor_id, so the current-owner check alone would
        # make unassignment un-undoable). A student meanwhile claimed by
        # another counselor matches neither arm and stays untouched — undo
        # must never pull a student off someone else's caseload.
        if not s:
            continue
        is_mine = s.assigned_counselor_id == current_user.id
        unowned_was_mine = (s.assigned_counselor_id in (None, 0)
                            and entry.get('assigned_counselor_id') == current_user.id)
        if not (is_mine or unowned_was_mine):
            continue
        rollover_util.restore(s, entry)
        log_action('rollover.undo', resource_type='student',
                   resource_id=s.id,
                   details=f'snapshot={snapshot.id} undone_action={entry.get("applied_action")}')
        restored += 1

    snapshot.undone = True
    snapshot.undone_at = datetime.now(timezone.utc)
    db.session.commit()
    log_action('rollover.undo_commit', resource_type='rollover_snapshot',
               resource_id=snapshot.id, details=f'restored={restored}')

    flash(f'Rollover undone. {restored} students restored.', 'success')
    return redirect(url_for('caseload.rollover'))
