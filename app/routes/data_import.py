"""Routes for importing attendance, grades, and other student data via CSV/Excel."""
import io
import csv
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from app import db
from app.models.student import Student
from app.models.attendance import AttendanceRecord
from app.models.grade import GradeRecord
from app.models.note import Note
from app.models.activity import Activity
from app.models.service_record import ServiceRecord
from app.utils.audit import log_action
from app.utils.helpers import parse_date
from datetime import date

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

data_import_bp = Blueprint('data_import', __name__)

VALID_ATTENDANCE = {'present', 'absent', 'tardy', 'excused'}
VALID_GRADES = {'A+', 'A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-',
                'D+', 'D', 'D-', 'F', 'P', 'NP', 'I', 'W'}

# Synergy SIS attendance code → (status, reason)
SYNERGY_STATUS_MAP = {
    '':                 ('present', ''),
    'activity':         ('excused', 'Activity'),
    'illness':          ('excused', 'Illness'),
    'excused':          ('excused', 'Excused'),
    'counseling':       ('excused', 'Counseling'),
    'testing':          ('excused', 'Testing'),
    'office excused':   ('excused', 'Office Excused'),
    'office ex':        ('excused', 'Office Excused'),
    'cut':              ('absent', 'Cut'),
    'unverified':       ('absent', 'Unverified'),
    'parent unexcused': ('absent', 'Parent Unexcused'),
    'tardy':            ('tardy', 'Tardy'),
    'unexcused tardy':  ('tardy', 'Unexcused Tardy'),
}


# =====================================================================
#  MAIN IMPORT HUB
# =====================================================================

@data_import_bp.route('/')
@login_required
def index():
    """Data import hub page."""
    # Count existing records
    student_ids = [row[0] for row in Student.query.filter_by(
        assigned_counselor_id=current_user.id).with_entities(Student.id).all()]
    attendance_count = AttendanceRecord.query.filter(
        AttendanceRecord.student_id.in_(student_ids)).count() if student_ids else 0
    grade_count = GradeRecord.query.filter(
        GradeRecord.student_id.in_(student_ids)).count() if student_ids else 0
    return render_template('data_import/index.html',
                           attendance_count=attendance_count,
                           grade_count=grade_count)


# =====================================================================
#  ATTENDANCE IMPORT
# =====================================================================

@data_import_bp.route('/attendance/template')
@login_required
def attendance_template():
    """Download attendance import Excel template."""
    if not HAS_OPENPYXL:
        flash('Excel support requires openpyxl. Install with: pip install openpyxl', 'danger')
        return redirect(url_for('data_import.index'))

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Import"

    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='2C5F8A', end_color='2C5F8A', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    columns = [
        ('Student ID #', 16),
        ('Date', 14),
        ('Period', 8),
        ('Status', 14),
        ('Course Name', 30),
        ('Reason', 30),
    ]

    for col_idx, (name, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data validation for Status
    status_dv = DataValidation(
        type='list', formula1='"Present,Absent,Tardy,Excused"', allow_blank=False,
        showErrorMessage=True, errorTitle='Invalid Status',
        error='Choose: Present, Absent, Tardy, or Excused'
    )
    status_dv.sqref = 'D2:D5000'
    ws.add_data_validation(status_dv)

    # Period validation (0-10 to support Synergy periods)
    period_dv = DataValidation(
        type='list', formula1='"0,1,2,3,4,5,6,7,8,9,10"', allow_blank=True,
        showErrorMessage=True, errorTitle='Invalid Period',
        error='Enter period 0-10'
    )
    period_dv.sqref = 'C2:C5000'
    ws.add_data_validation(period_dv)

    # Instructions sheet
    instr = wb.create_sheet('Instructions')
    instr.sheet_properties.tabColor = 'E8A838'
    instructions = [
        ('ATTENDANCE IMPORT TEMPLATE', ''),
        ('', ''),
        ('Column', 'Instructions'),
        ('Student ID #', 'Required. Must match a student in your caseload.'),
        ('Date', 'Required. Format: MM/DD/YYYY or YYYY-MM-DD'),
        ('Period', 'Optional. Class period 0-10. (1-4 core, 5 extracurricular, 6 advisory)'),
        ('Status', 'Required. Present, Absent, Tardy, or Excused.'),
        ('Course Name', 'Optional. Name of the class.'),
        ('Reason', 'Optional. Reason for absence/tardy.'),
        ('', ''),
        ('TIPS:', ''),
        ('', 'You can paste data exported from your SIS (Synergy, Aeries, PowerSchool, etc.)'),
        ('', 'One row per student per period. Daily attendance = one row per student.'),
        ('', 'Duplicate rows (same student+date+period) will be skipped.'),
    ]
    for row_idx, (a, b) in enumerate(instructions, 1):
        instr.cell(row=row_idx, column=1, value=a).font = Font(
            name='Calibri', bold=bool(a), size=14 if row_idx == 1 else 11,
            color='2C5F8A' if row_idx == 1 else '000000')
        instr.cell(row=row_idx, column=2, value=b).font = Font(name='Calibri', size=11)
    instr.column_dimensions['A'].width = 18
    instr.column_dimensions['B'].width = 75

    ws.freeze_panes = 'A2'
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    log_action('export', 'attendance_template', details='Downloaded attendance template')
    return send_file(buffer,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='Attendance_Import_Template.xlsx')


@data_import_bp.route('/attendance/upload', methods=['GET', 'POST'])
@login_required
def attendance_upload():
    """Upload attendance data from Excel or CSV."""
    if request.method == 'POST':
        file = request.files.get('file')
        if not file:
            flash('Please select a file.', 'danger')
            return redirect(url_for('data_import.attendance_upload'))

        header, rows = _parse_upload_file(file)
        if header is None:
            return redirect(url_for('data_import.attendance_upload'))

        # Auto-detect Synergy format and convert to standard rows
        is_synergy = _is_synergy_format(header)
        if is_synergy:
            rows = _convert_synergy_rows(header, rows)
            if rows is None:
                flash('Could not parse Synergy file. Check that it has Perm ID, Date, and Period columns.', 'danger')
                return redirect(url_for('data_import.attendance_upload'))

        added = 0
        skipped = 0
        not_on_caseload = 0
        errors = []

        # Pre-load existing attendance keys to avoid per-row duplicate checks
        caseload_ids = list(student_cache.values())
        existing_keys = set()
        if caseload_ids:
            for sid, att_date, period in db.session.query(
                AttendanceRecord.student_id, AttendanceRecord.date, AttendanceRecord.period
            ).filter(AttendanceRecord.student_id.in_(caseload_ids)).all():
                existing_keys.add((sid, att_date, period))

        for row_idx, row in enumerate(rows, start=2):
            if len(row) < 4:
                row.extend([''] * (6 - len(row)))
            student_id_str, date_str, period_str, status_str = row[0], row[1], row[2], row[3]
            course_name = row[4] if len(row) > 4 else ''
            reason = row[5] if len(row) > 5 else ''

            if not student_id_str and not date_str and not status_str:
                continue

            row_errors = []
            if not student_id_str:
                row_errors.append('Student ID # required')
            if not date_str:
                row_errors.append('Date required')
            if not status_str:
                row_errors.append('Status required')

            # Synergy rows already have clean status; template rows need validation
            status_clean = str(status_str or '').strip().lower()
            if status_clean and status_clean not in VALID_ATTENDANCE:
                row_errors.append(f'Invalid status: {status_str}')

            att_date = parse_date(str(date_str).strip()) if date_str else None
            if date_str and not att_date:
                row_errors.append(f'Invalid date: {date_str}')

            period_val = None
            if period_str:
                try:
                    period_val = int(float(period_str))
                    if period_val < 0 or period_val > 10:
                        row_errors.append(f'Period must be 0-10, got {period_val}')
                except (ValueError, TypeError):
                    row_errors.append(f'Invalid period: {period_str}')

            if row_errors:
                if not is_synergy:
                    errors.append(f'Row {row_idx}: ' + '; '.join(row_errors))
                continue

            student = Student.query.filter_by(
                student_id_number=str(student_id_str).strip()).first()
            if not student:
                # For Synergy whole-school reports, silently skip non-caseload students
                if is_synergy:
                    not_on_caseload += 1
                    continue
                errors.append(f'Row {row_idx}: Student ID {student_id_str} not found')
                continue

            # Skip duplicates (check in-memory set instead of DB query)
            att_key = (student.id, att_date, period_val)
            if att_key in existing_keys:
                skipped += 1
                continue
            existing_keys.add(att_key)  # Track newly added records too

            record = AttendanceRecord(
                student_id=student.id,
                date=att_date,
                period=period_val,
                status=status_clean,
                course_name=str(course_name or '').strip(),
                reason=str(reason or '').strip(),
                imported_by_id=current_user.id,
            )
            db.session.add(record)
            added += 1

        db.session.commit()
        log_action('import', 'attendance',
                   details=f'Imported attendance: {added} added, {skipped} skipped'
                           + (f', {not_on_caseload} not on caseload' if not_on_caseload else ''))

        if is_synergy:
            fmt_msg = f'Synergy import: {added} records added, {skipped} duplicates skipped.'
            if not_on_caseload:
                fmt_msg += f' {not_on_caseload} records skipped (students not on your caseload).'
            if errors:
                fmt_msg += f' {len(errors)} errors.'
                flash(fmt_msg, 'warning')
            else:
                flash(fmt_msg, 'success')
        else:
            if errors:
                flash(f'Imported with issues: {added} added, {skipped} duplicates skipped, {len(errors)} errors.', 'warning')
            else:
                flash(f'Attendance imported: {added} records added, {skipped} duplicates skipped.', 'success')

        return render_template('data_import/attendance_upload.html',
                               added=added, skipped=skipped, errors=errors)

    return render_template('data_import/attendance_upload.html',
                           added=0, skipped=0, errors=None)


# =====================================================================
#  GRADES IMPORT
# =====================================================================

@data_import_bp.route('/grades/template')
@login_required
def grades_template():
    """Download grades import Excel template."""
    if not HAS_OPENPYXL:
        flash('Excel support requires openpyxl.', 'danger')
        return redirect(url_for('data_import.index'))

    wb = Workbook()
    ws = wb.active
    ws.title = "Grades Import"

    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='2C5F8A', end_color='2C5F8A', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    columns = [
        ('School Year', 14),
        ('Perm ID', 12),
        ('Grade Level', 12),
        ('Grade', 8),
        ('Mark Order', 12),
        ('Mark Name', 14),
        ('Course Title', 30),
        ('Course ID', 12),
        ('Period', 8),
        ('C1', 6),
        ('C2', 6),
        ('C3', 6),
        ('Staff Name', 20),
        ('Audit Class', 12),
        ('Student Name', 22),
        ('SPED', 8),
        ('SigDis', 8),
        ('SLE', 8),
    ]

    for col_idx, (name, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Validations
    grade_dv = DataValidation(
        type='list',
        formula1='"A+,A,A-,B+,B,B-,C+,C,C-,D+,D,D-,F,P,NP,I,W"',
        allow_blank=True, showErrorMessage=True, errorTitle='Invalid Grade',
        error='Enter a valid letter grade')
    grade_dv.sqref = 'D2:D5000'
    ws.add_data_validation(grade_dv)

    audit_dv = DataValidation(
        type='list', formula1='"N,Y"', allow_blank=True)
    audit_dv.sqref = 'N2:N5000'
    ws.add_data_validation(audit_dv)

    # Instructions
    instr = wb.create_sheet('Instructions')
    instr.sheet_properties.tabColor = 'E8A838'
    instructions = [
        ('GRADES IMPORT TEMPLATE — SYNERGY FORMAT', ''),
        ('', ''),
        ('This template matches the Synergy SIS grade export.', ''),
        ('Export from Synergy and paste or upload directly.', ''),
        ('', ''),
        ('Column', 'Instructions'),
        ('School Year', 'e.g., 2025-2026.'),
        ('Perm ID', 'Required. Student permanent ID — must match a student in your caseload.'),
        ('Grade Level', 'Student grade level (9, 10, 11, 12). Informational.'),
        ('Grade', 'Required. Letter grade (A+, A, A-, B+, … F, P, NP, I, W).'),
        ('Mark Order', 'Numeric mark sequence (1, 2, etc.). Optional.'),
        ('Mark Name', 'e.g., "Quarter 3". The quarter number is extracted automatically.'),
        ('Course Title', 'Required. Full course title from Synergy (e.g., "Spanish 1 CP [S1]").'),
        ('Course ID', 'Synergy course ID number. Used to match to your Course Catalog.'),
        ('Period', 'Class period (1-4).'),
        ('C1, C2, C3', 'Citizenship grades. Stored for reference but not analyzed.'),
        ('Staff Name', 'Teacher name. Stored for reference.'),
        ('Audit Class', 'N or Y. Audit classes are skipped during import.'),
        ('Student Name', 'For reference. Not used for matching (Perm ID is used).'),
        ('SPED', 'Special education flag. Informational.'),
        ('SigDis', 'Significant disability flag. Informational.'),
        ('SLE', 'Structured Learning Experience flag. Informational.'),
        ('', ''),
        ('TIPS:', ''),
        ('', 'You can paste your Synergy export directly — no reformatting needed.'),
        ('', 'One row per student per course per quarter.'),
        ('', 'The AI will use this data to recommend courses for next year.'),
    ]
    for row_idx, (a, b) in enumerate(instructions, 1):
        instr.cell(row=row_idx, column=1, value=a).font = Font(
            name='Calibri', bold=bool(a), size=14 if row_idx == 1 else 11,
            color='2C5F8A' if row_idx == 1 else '000000')
        instr.cell(row=row_idx, column=2, value=b).font = Font(name='Calibri', size=11)
    instr.column_dimensions['A'].width = 18
    instr.column_dimensions['B'].width = 75

    ws.freeze_panes = 'A2'
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    log_action('export', 'grades_template', details='Downloaded grades template')
    return send_file(buffer,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='Grades_Import_Template.xlsx')


@data_import_bp.route('/grades/upload', methods=['GET', 'POST'])
@login_required
def grades_upload():
    """Upload grades data from Excel or CSV."""
    if request.method == 'POST':
        file = request.files.get('file')
        if not file:
            flash('Please select a file.', 'danger')
            return redirect(url_for('data_import.grades_upload'))

        _header, rows = _parse_upload_file(file)
        if rows is None:
            return redirect(url_for('data_import.grades_upload'))

        added = 0
        updated = 0
        skipped = 0
        errors = []
        BATCH_SIZE = 200

        # Pre-load student lookup cache
        student_cache = {
            s.student_id_number: s.id
            for s in Student.query.with_entities(
                Student.student_id_number, Student.id).all()
        }

        # Build column index from header row
        col_map = _build_grade_col_map(_header)

        # Quarter may come from the header name (e.g., "Quarter 3" column)
        header_quarter = col_map.pop('_quarter_from_header', None)

        for row_idx, row in enumerate(rows, start=2):
            # Pad short rows
            while len(row) < 16:
                row.append('')

            # Extract values by mapped column positions
            student_id_str = _col(row, col_map, 'perm_id')
            school_year = _col(row, col_map, 'school_year')
            letter_grade = _col(row, col_map, 'grade')
            mark_name = _col(row, col_map, 'mark_name')
            course_name = _col(row, col_map, 'course_title')
            course_number = _col(row, col_map, 'course_id')
            period_str = _col(row, col_map, 'period')
            audit_class = _col(row, col_map, 'audit_class')
            credits_att = _col(row, col_map, 'credits_att')

            if not student_id_str and not course_name:
                continue

            # Skip audit classes
            if str(audit_class).strip().upper() == 'Y':
                skipped += 1
                continue

            row_errors = []
            if not student_id_str:
                row_errors.append('Perm ID required')
            if not course_name:
                row_errors.append('Course Title required')

            # ── Determine quarter ──
            # Priority: header name ("Quarter 3" column) > mark_name field > None
            quarter_val = header_quarter or _parse_quarter(mark_name)

            # Derive semester from quarter (Q1-Q2 = Sem 1, Q3-Q4 = Sem 2)
            semester_val = 2 if quarter_val and quarter_val >= 3 else 1

            # ── Parse period ──
            # Synergy grade reports may have Period column or Section ID like "1-010"
            period_val = None
            if period_str:
                try:
                    period_val = int(float(period_str))
                except (ValueError, TypeError):
                    pass
            # If no Period column but Section ID has "1-010" format, extract period
            if period_val is None and course_number and '-' in str(course_number):
                try:
                    period_val = int(str(course_number).split('-')[0])
                except (ValueError, TypeError):
                    pass

            # ── Parse per-course credits ──
            credits_val = 5.0
            if credits_att:
                try:
                    credits_val = float(credits_att)
                except (ValueError, TypeError):
                    pass

            # Validate letter grade
            letter_clean = str(letter_grade or '').strip()
            if letter_clean and letter_clean not in VALID_GRADES:
                row_errors.append(f'Invalid grade: {letter_clean}')

            if row_errors:
                errors.append(f'Row {row_idx}: ' + '; '.join(row_errors))
                continue

            # Skip rows with no letter grade (empty grade cell)
            if not letter_clean:
                skipped += 1
                continue

            sid_clean = str(student_id_str).strip()
            student_db_id = student_cache.get(sid_clean)
            if not student_db_id:
                errors.append(f'Row {row_idx}: Perm ID {student_id_str} not found')
                continue

            # Detect honors/AP from course title
            course_title_clean = str(course_name).strip()
            title_upper = course_title_clean.upper()
            is_honors_ap = 'AP ' in title_upper or title_upper.startswith('AP ') or ' HONORS' in title_upper

            # Upsert: update if same student+year+quarter+course
            school_year_clean = str(school_year or '').strip()
            existing = GradeRecord.query.filter_by(
                student_id=student_db_id,
                school_year=school_year_clean,
                quarter=quarter_val,
                course_name=course_title_clean,
            ).first()

            if existing:
                existing.letter_grade = letter_clean or existing.letter_grade
                existing.course_number = str(course_number or '').strip() or existing.course_number
                existing.is_honors_ap = is_honors_ap
                existing.credits_earned = credits_val
                updated += 1
            else:
                record = GradeRecord(
                    student_id=student_db_id,
                    school_year=school_year_clean,
                    quarter=quarter_val,
                    course_name=course_title_clean,
                    course_number=str(course_number or '').strip(),
                    period=period_val,
                    letter_grade=letter_clean,
                    credits_earned=credits_val,
                    credits_attempted=credits_val,
                    is_semester=semester_val,
                    is_honors_ap=is_honors_ap,
                    imported_by_id=current_user.id,
                )
                db.session.add(record)
                added += 1

            # Batch commit to avoid holding SQLite lock too long
            if (added + updated) % BATCH_SIZE == 0:
                db.session.commit()

        db.session.commit()
        log_action('import', 'grades',
                   details=f'Imported grades: {added} added, {updated} updated')

        if errors:
            flash(f'Imported with issues: {added} added, {updated} updated, {len(errors)} errors.', 'warning')
        else:
            flash(f'Grades imported: {added} added, {updated} updated.', 'success')

        return render_template('data_import/grades_upload.html',
                               added=added, updated=updated, errors=errors)

    return render_template('data_import/grades_upload.html',
                           added=0, updated=0, errors=None)


# =====================================================================
#  STUDENT INFO BULK UPDATE
# =====================================================================

# Fields that can be bulk-updated via spreadsheet
STUDENT_UPDATE_FIELDS = {
    'first_name':              ('First Name', str),
    'last_name':               ('Last Name', str),
    'grade_level':             ('Grade Level', int),
    'gender':                  ('Gender', str),
    'ethnicity':               ('Ethnicity', str),
    'email':                   ('Student Email', str),
    'phone':                   ('Phone', str),
    'date_of_birth':           ('Date of Birth', 'date'),
    'parent_guardian_name':    ('Parent/Guardian', str),
    'parent_guardian_phone':   ('Parent Phone', str),
    'parent_guardian_email':   ('Parent Email', str),
    'address':                 ('Address', str),
    'homeroom':                ('Advisory', str),
    'el_status':               ('EL Status', str),
    'el_level':                ('EL Level', str),
    'iep_status':              ('IEP', 'bool'),
    'section_504':             ('504', 'bool'),
}


@data_import_bp.route('/students/template')
@login_required
def student_update_template():
    """Download a pre-filled Excel template with current student data for bulk editing."""
    if not HAS_OPENPYXL:
        flash('Excel support requires openpyxl. Install it with: pip install openpyxl', 'danger')
        return redirect(url_for('data_import.index'))

    wb = Workbook()
    ws = wb.active
    ws.title = "Student Update"

    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2C5F8A', end_color='2C5F8A', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    # Column order: Student ID (key), then all updatable fields
    columns = [
        ('Student ID #', 18),
        ('First Name', 16),
        ('Last Name', 16),
        ('Grade Level', 12),
        ('Gender', 14),
        ('Ethnicity', 16),
        ('Date of Birth', 14),
        ('Advisory', 14),
        ('Student Email', 26),
        ('Phone', 16),
        ('Parent/Guardian', 22),
        ('Parent Phone', 16),
        ('Parent Email', 26),
        ('Address', 30),
        ('EL Status', 14),
        ('EL Level', 10),
        ('IEP', 6),
        ('504', 6),
    ]

    # Map column header to db field name
    HEADER_TO_FIELD = {label: field for field, (label, _) in STUDENT_UPDATE_FIELDS.items()}
    HEADER_TO_FIELD['Student ID #'] = 'student_id_number'

    for col_idx, (name, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Pre-fill with current student data from caseload
    students = Student.query.filter_by(
        assigned_counselor_id=current_user.id
    ).filter(Student.status == 'active').order_by(Student.last_name, Student.first_name).all()

    for row_idx, s in enumerate(students, start=2):
        row_data = [
            s.student_id_number,
            s.first_name,
            s.last_name,
            s.grade_level,
            s.gender or '',
            s.ethnicity or '',
            s.date_of_birth.strftime('%m/%d/%Y') if s.date_of_birth else '',
            s.homeroom or '',
            s.email or '',
            s.phone or '',
            s.parent_guardian_name or '',
            s.parent_guardian_phone or '',
            s.parent_guardian_email or '',
            s.address or '',
            s.el_status or '',
            s.el_level or '',
            'Yes' if s.iep_status else '',
            'Yes' if s.section_504 else '',
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    # Lock the Student ID column (light gray background to indicate read-only)
    lock_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    for row_idx in range(2, len(students) + 2):
        ws.cell(row=row_idx, column=1).fill = lock_fill

    # Instructions sheet
    instr = wb.create_sheet('Instructions')
    instructions = [
        ('Student Info Bulk Update', ''),
        ('', ''),
        ('How it works', 'Edit any cells in the Student Update sheet, then upload the file. '
                         'Students are matched by Student ID #. Only changed fields are updated.'),
        ('Student ID #', 'DO NOT change this column — it is the key used to match students.'),
        ('Adding new columns', 'You can delete columns you do not need. '
                               'Only columns with recognized headers will be processed.'),
        ('Blank cells', 'Blank cells are skipped (existing value is kept). '
                        'To clear a field, enter a single dash (-).'),
        ('IEP / 504', 'Enter "Yes" to enable, "No" to disable, or leave blank to keep current value.'),
        ('EL Status', 'Valid values: EO, Newcomer, LTEL, RFEP'),
        ('Date of Birth', 'Format: MM/DD/YYYY or YYYY-MM-DD'),
        ('Grade Level', 'Enter the number only (e.g. 9, 10, 11, 12)'),
    ]
    for row_idx, (a, b) in enumerate(instructions, 1):
        instr.cell(row=row_idx, column=1, value=a).font = Font(
            name='Calibri', bold=bool(a), size=14 if row_idx == 1 else 11,
            color='2C5F8A' if row_idx == 1 else '000000')
        instr.cell(row=row_idx, column=2, value=b).font = Font(name='Calibri', size=11)
    instr.column_dimensions['A'].width = 20
    instr.column_dimensions['B'].width = 80

    ws.freeze_panes = 'A2'
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    log_action('export', 'student_update_template', details='Downloaded student update template')
    return send_file(buffer,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='Student_Info_Update.xlsx')


@data_import_bp.route('/students/upload', methods=['GET', 'POST'])
@login_required
def student_update_upload():
    """Bulk update student profiles from an uploaded spreadsheet."""
    if request.method == 'POST':
        file = request.files.get('file')
        if not file:
            flash('Please select a file.', 'danger')
            return redirect(url_for('data_import.student_update_upload'))

        header, rows = _parse_upload_file(file)
        if header is None:
            return redirect(url_for('data_import.student_update_upload'))

        # Build header → db field mapping from the uploaded column names
        header_lower = [h.strip().lower() for h in header]
        LABEL_TO_FIELD = {label.lower(): field for field, (label, _) in STUDENT_UPDATE_FIELDS.items()}
        LABEL_TO_FIELD['student id #'] = 'student_id_number'
        LABEL_TO_FIELD['student id'] = 'student_id_number'
        LABEL_TO_FIELD['perm id'] = 'student_id_number'
        LABEL_TO_FIELD['advisory'] = 'homeroom'
        LABEL_TO_FIELD['homeroom'] = 'homeroom'
        LABEL_TO_FIELD['dob'] = 'date_of_birth'

        col_map = {}  # col_index → field_name
        id_col = None
        for idx, h in enumerate(header_lower):
            field = LABEL_TO_FIELD.get(h)
            if field == 'student_id_number':
                id_col = idx
            elif field and field in STUDENT_UPDATE_FIELDS:
                col_map[idx] = field

        if id_col is None:
            flash('Could not find a Student ID column. Expected "Student ID #", "Student ID", or "Perm ID".', 'danger')
            return redirect(url_for('data_import.student_update_upload'))

        if not col_map:
            flash('No recognized data columns found. Check that your column headers match the template.', 'danger')
            return redirect(url_for('data_import.student_update_upload'))

        # Pre-load students on caseload
        student_cache = {}
        for s in Student.query.filter_by(assigned_counselor_id=current_user.id).all():
            student_cache[s.student_id_number] = s

        updated = 0
        skipped = 0
        not_found = 0
        field_changes = 0
        errors = []
        BATCH_SIZE = 200

        for row_idx, row in enumerate(rows, start=2):
            if len(row) <= id_col:
                continue
            sid = str(row[id_col]).strip()
            if not sid:
                continue

            student = student_cache.get(sid)
            if not student:
                not_found += 1
                continue

            changed = False
            for col_idx, field in col_map.items():
                if col_idx >= len(row):
                    continue
                raw = str(row[col_idx]).strip() if row[col_idx] else ''
                if not raw:
                    continue  # blank = keep existing

                # Clear field with dash
                if raw == '-':
                    if getattr(student, field):
                        setattr(student, field, '' if isinstance(getattr(student, field), str) else None)
                        changed = True
                        field_changes += 1
                    continue

                _, ftype = STUDENT_UPDATE_FIELDS[field]

                try:
                    if ftype == 'bool':
                        new_val = raw.lower() in ('yes', 'true', '1', 'y')
                        if raw.lower() in ('no', 'false', '0', 'n'):
                            new_val = False
                        elif raw.lower() not in ('yes', 'true', '1', 'y'):
                            continue  # unrecognized boolean, skip
                        if getattr(student, field) != new_val:
                            setattr(student, field, new_val)
                            changed = True
                            field_changes += 1
                    elif ftype == 'date':
                        new_date = parse_date(raw)
                        if new_date and getattr(student, field) != new_date:
                            setattr(student, field, new_date)
                            changed = True
                            field_changes += 1
                        elif not new_date:
                            errors.append(f'Row {row_idx}: Invalid date "{raw}" for {field}')
                    elif ftype == int:
                        new_val = int(float(raw))
                        if getattr(student, field) != new_val:
                            setattr(student, field, new_val)
                            changed = True
                            field_changes += 1
                    else:  # str
                        if getattr(student, field) != raw:
                            setattr(student, field, raw)
                            changed = True
                            field_changes += 1
                except (ValueError, TypeError) as e:
                    errors.append(f'Row {row_idx}: Bad value "{raw}" for {field}: {e}')
                    continue

            if changed:
                updated += 1
            else:
                skipped += 1

            if updated > 0 and updated % BATCH_SIZE == 0:
                db.session.commit()

        db.session.commit()
        log_action('import', 'student_update',
                   details=f'Bulk student update: {updated} students updated, '
                           f'{field_changes} fields changed, {skipped} unchanged')

        if errors:
            flash(f'Updated {updated} students ({field_changes} fields changed), '
                  f'{skipped} unchanged, {not_found} not on caseload, {len(errors)} errors.', 'warning')
        else:
            msg = f'Updated {updated} students ({field_changes} fields changed).'
            if skipped:
                msg += f' {skipped} already up to date.'
            if not_found:
                msg += f' {not_found} student IDs not on your caseload.'
            flash(msg, 'success')

        return render_template('data_import/student_update_upload.html',
                               updated=updated, skipped=skipped, errors=errors)

    return render_template('data_import/student_update_upload.html',
                           updated=0, skipped=0, errors=None)


# =====================================================================
#  CLEAR DATA
# =====================================================================

@data_import_bp.route('/attendance/clear', methods=['POST'])
@login_required
def clear_attendance():
    """Clear all attendance records for current counselor's students."""
    student_ids = [row[0] for row in Student.query.filter_by(
        assigned_counselor_id=current_user.id).with_entities(Student.id).all()]
    count = AttendanceRecord.query.filter(
        AttendanceRecord.student_id.in_(student_ids)).delete(synchronize_session=False)
    db.session.commit()
    log_action('delete', 'attendance', details=f'Cleared {count} attendance records')
    flash(f'Cleared {count} attendance records.', 'warning')
    return redirect(url_for('data_import.index'))


@data_import_bp.route('/grades/clear', methods=['POST'])
@login_required
def clear_grades():
    """Clear all grade records for current counselor's students."""
    student_ids = [row[0] for row in Student.query.filter_by(
        assigned_counselor_id=current_user.id).with_entities(Student.id).all()]
    count = GradeRecord.query.filter(
        GradeRecord.student_id.in_(student_ids)).delete(synchronize_session=False)
    db.session.commit()
    log_action('delete', 'grades', details=f'Cleared {count} grade records')
    flash(f'Cleared {count} grade records.', 'warning')
    return redirect(url_for('data_import.index'))


# =====================================================================
#  NOTES → SESSIONS CONVERSION
# =====================================================================

# Note type → Activity (service_type, category, delivery_type)
_NOTE_TO_ACTIVITY = {
    'individual':      ('direct_student', 'individual_counseling', 'individual'),
    'group':           ('direct_student', 'group_counseling', 'small_group'),
    'parent_contact':  ('indirect_student', 'parent_outreach', 'individual'),
    'teacher_consult': ('indirect_student', 'consultation', 'individual'),
    'crisis':          ('direct_student', 'crisis_response', 'individual'),
    'follow_up':       ('direct_student', 'individual_counseling', 'individual'),
    'referral':        ('indirect_student', 'referrals', 'individual'),
    'observation':     ('direct_student', 'appraisal', 'individual'),
    'classroom':       ('direct_student', 'classroom_instruction', 'classroom'),
    'college_career':  ('direct_student', 'advisement', 'individual'),
    'assessment':      ('direct_student', 'appraisal', 'individual'),
}

# Note type → ServiceRecord service_type
_NOTE_TO_SERVICE = {
    'individual':      'individual_counseling',
    'group':           'group_counseling',
    'parent_contact':  'parent_conference',
    'teacher_consult': 'consultation',
    'crisis':          'crisis_intervention',
    'follow_up':       'follow_up',
    'referral':        'referral',
    'observation':     'observation',
    'classroom':       'classroom_lesson',
    'college_career':  'college_career',
    'assessment':      'assessment',
}

# Note delivery_method → ServiceRecord setting
_DELIVERY_TO_SETTING = {
    'in_person': 'office',
    'phone':     'phone',
    'email':     'email',
    'virtual':   'virtual',
}


@data_import_bp.route('/convert-notes', methods=['GET', 'POST'])
@login_required
def convert_notes():
    """Convert existing Counseling Notes into Activity Log + Service Record entries."""
    # Count notes that haven't been converted yet
    my_notes = Note.query.filter_by(author_id=current_user.id).all()
    total_notes = len(my_notes)

    if request.method == 'GET':
        return render_template('data_import/convert_notes.html',
                               total_notes=total_notes,
                               converted=None, skipped=None, errors=None)

    # POST — run the conversion
    activities_created = 0
    services_created = 0
    skipped = 0
    errors = []

    for note in my_notes:
        try:
            # --- Create Activity entry (use-of-time) ---
            svc_type, category, delivery_type = _NOTE_TO_ACTIVITY.get(
                note.note_type, ('direct_student', 'individual_counseling', 'individual'))

            # Check for duplicate activity (same counselor + date + title)
            title = note.title or f'{note.note_type.replace("_", " ").title()} Session'
            existing_activity = Activity.query.filter_by(
                counselor_id=current_user.id,
                date=note.session_date,
                title=title,
            ).first()

            if existing_activity:
                skipped += 1
                continue

            activity = Activity(
                counselor_id=current_user.id,
                title=title,
                description=note.content[:500] if note.content else None,
                date=note.session_date,
                duration_minutes=note.duration_minutes or 30,
                service_type=svc_type,
                category=category,
                topic=note.topic_category,
                delivery_type=delivery_type,
                num_students=1 if note.note_type != 'classroom' else 0,
            )
            db.session.add(activity)
            activities_created += 1

            # --- Create ServiceRecord (per-student service log) ---
            svc_record_type = _NOTE_TO_SERVICE.get(
                note.note_type, 'individual_counseling')

            existing_service = ServiceRecord.query.filter_by(
                student_id=note.student_id,
                counselor_id=current_user.id,
                date=note.session_date,
                service_type=svc_record_type,
                topic=note.topic_category,
            ).first()

            if not existing_service:
                service = ServiceRecord(
                    student_id=note.student_id,
                    counselor_id=current_user.id,
                    date=note.session_date,
                    service_type=svc_record_type,
                    topic=note.topic_category,
                    description=note.content[:500] if note.content else None,
                    duration_minutes=note.duration_minutes or 30,
                    asca_domain=note.asca_domain,
                    delivery_method=note.delivery_method,
                    setting=_DELIVERY_TO_SETTING.get(note.delivery_method, 'office'),
                    follow_up_required=note.follow_up_needed,
                    follow_up_date=note.follow_up_date,
                    referral_made=(note.note_type == 'referral'),
                )
                db.session.add(service)
                services_created += 1

        except Exception as e:
            errors.append(f'Note #{note.id}: {str(e)}')

    db.session.commit()
    log_action('convert', 'notes_to_sessions',
               details=f'Converted notes: {activities_created} activities, {services_created} service records, {skipped} skipped')

    if errors:
        flash(f'Converted with issues: {activities_created} activities, {services_created} service records, {skipped} skipped, {len(errors)} errors.', 'warning')
    else:
        flash(f'Converted {activities_created} activities and {services_created} service records from {total_notes} notes. {skipped} duplicates skipped.', 'success')

    return render_template('data_import/convert_notes.html',
                           total_notes=total_notes,
                           converted={'activities': activities_created, 'services': services_created},
                           skipped=skipped, errors=errors)


# =====================================================================
#  HELPERS
# =====================================================================

def _parse_upload_file(file):
    """Parse CSV or Excel file, return (header_row, data_rows).

    Returns (None, None) on error.  header_row is a list of strings
    (the first row), data_rows is a list-of-lists for the remaining rows.
    """
    filename = file.filename.lower()

    if filename.endswith('.csv'):
        try:
            text = file.read().decode('utf-8-sig')
            reader = csv.reader(text.splitlines())
            rows = list(reader)
            if rows:
                return rows[0], rows[1:]
            return [], []
        except Exception as e:
            flash(f'Could not read CSV: {str(e)}', 'danger')
            return None, None

    elif filename.endswith(('.xlsx', '.xls')):
        if not HAS_OPENPYXL:
            flash('Excel support requires openpyxl.', 'danger')
            return None, None
        try:
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            header = []
            rows = []
            for idx, row in enumerate(ws.iter_rows(values_only=True)):
                str_row = [str(c) if c is not None else '' for c in row]
                if idx == 0:
                    header = str_row
                else:
                    rows.append(str_row)
            return header, rows
        except Exception as e:
            flash(f'Could not read Excel file: {str(e)}', 'danger')
            return None, None
    else:
        flash('Please upload a .csv or .xlsx file.', 'danger')
        return None, None


import re as _re


# ── Grade import helpers ─────────────────────────────────────────

# Canonical header names → accepted header variations (case-insensitive)
_GRADE_COL_ALIASES = {
    'school_year':  ('school year', 'schoolyear', 'year'),
    'perm_id':      ('perm id', 'permid', 'student id', 'student id #', 'student_id'),
    'grade_level':  ('grade level', 'gradelevel', 'grd', 'grade'),
    'grade':        ('letter grade', 'lettergrade', 'mark'),
    'mark_order':   ('mark order', 'markorder'),
    'mark_name':    ('mark name', 'markname', 'term'),
    'course_title': ('course title', 'coursetitle', 'course name', 'coursename'),
    'course_id':    ('course id', 'courseid', 'course number', 'coursenumber', 'course #', 'section id', 'sectionid'),
    'period':       ('period', 'per'),
    'audit_class':  ('audit class', 'auditclass', 'audit'),
    'staff_name':   ('staff name', 'staffname', 'teacher', 'teacher name'),
    'student_name': ('student name', 'studentname', 'name'),
    'credits_att':  ('credits att', 'credits attempted', 'cred att', 'credits'),
    'credits_comp': ('credits completed', 'cred comp', 'credits comp'),
    'gpa':          ('gpa',),
    'gender':       ('gender',),
}


def _build_grade_col_map(header):
    """Map canonical column names to 0-based indices from the actual header row.

    Handles Synergy grade report format where the letter grade column is named
    "Quarter 3" (or "Quarter 1", etc.) — the header itself encodes the quarter.
    """
    if not header:
        return {}
    col_map = {}
    header_lower = [h.strip().lower() for h in header]

    for canon, aliases in _GRADE_COL_ALIASES.items():
        for alias in aliases:
            if alias in header_lower:
                col_map[canon] = header_lower.index(alias)
                break

    # ── Detect "Quarter X" column as the letter grade source ──
    # In Synergy grade reports, there's no separate "letter grade" or "mark name"
    # column. Instead, the header itself is "Quarter 3" and values are B, C-, etc.
    if 'grade' not in col_map:
        for i, h in enumerate(header_lower):
            m = _re.match(r'(quarter|qtr|q)\s*(\d)', h)
            if m:
                col_map['grade'] = i
                col_map['_quarter_from_header'] = int(m.group(2))
                break

    # ── Disambiguate "Grade" column ──
    # If "grade" and "grade_level" both mapped to the same column (because the
    # header just says "Grade"), check if values look like grade levels (9-12)
    # or letter grades (A, B, C). Prefer treating standalone "Grade" next to
    # "Perm ID" as grade level when a Quarter column exists for letter grades.
    if 'grade' in col_map and 'grade_level' in col_map and col_map['grade'] == col_map['grade_level']:
        # Both matched the same column — resolve ambiguity
        if '_quarter_from_header' in col_map:
            # We found a Quarter column for letter grades, so "Grade" = grade level
            del col_map['grade']  # remove; letter grade comes from Quarter col
        else:
            # No Quarter column found; look for another "Grade" column
            for i, h in enumerate(header_lower):
                if h == 'grade' and i != col_map['grade_level']:
                    col_map['grade'] = i
                    break

    return col_map


def _col(row, col_map, key):
    """Safely get a column value from a row by mapped key."""
    idx = col_map.get(key)
    if idx is None or idx >= len(row):
        return ''
    return str(row[idx]).strip()


def _parse_quarter(mark_name_str):
    """Extract quarter number from mark name like 'Quarter 3' or 'Q3'."""
    s = str(mark_name_str or '').strip()
    if not s:
        return None
    # Try "Quarter 3", "Q3", "Qtr 3", or just "3"
    m = _re.search(r'(?:quarter|qtr|q)\s*(\d)', s, _re.IGNORECASE)
    if m:
        return int(m.group(1))
    # Try bare number
    try:
        val = int(s)
        if 1 <= val <= 4:
            return val
    except (ValueError, TypeError):
        pass
    return None


def _is_synergy_format(header):
    """Detect if the header row matches a Synergy attendance export."""
    if not header:
        return False
    header_lower = [h.strip().lower() for h in header]
    # Synergy reports have "period 0", "period 1", … as column headers
    return 'period 0' in header_lower or ('period 1' in header_lower and 'perm id' in header_lower)


def _convert_synergy_rows(header, rows):
    """Convert Synergy pivot-format rows into standard flat attendance rows.

    Synergy format: Student Name | Perm ID | Grd | Date | Period 0 | Period 1 | … | Period N | Relation cols…
    Output rows:    [student_id, date, period, status, course_name, reason]

    Student name/ID/grade may be blank on continuation rows — they carry
    forward from the most recent row that had them.
    """
    header_lower = [h.strip().lower() for h in header]

    # Find column indices
    period_cols = {}  # period_number -> column_index
    for idx, h in enumerate(header_lower):
        if h.startswith('period '):
            try:
                period_num = int(h.split(' ', 1)[1])
                period_cols[period_num] = idx
            except (ValueError, IndexError):
                pass

    # Find key columns by name
    def find_col(names):
        for name in names:
            if name in header_lower:
                return header_lower.index(name)
        return None

    id_col = find_col(['perm id', 'student id', 'student id #'])
    date_col = find_col(['date'])
    name_col = find_col(['student name', 'name'])
    grade_col = find_col(['grd', 'grade'])

    if id_col is None or date_col is None or not period_cols:
        return None  # Not a valid Synergy file

    flat_rows = []
    # Carry-forward state for grouped student rows
    current_id = ''
    current_name = ''

    for row in rows:
        # Pad short rows
        if len(row) < len(header):
            row.extend([''] * (len(header) - len(row)))

        # Carry forward student info when blank
        row_id = row[id_col].strip() if id_col is not None else ''
        row_name = row[name_col].strip() if name_col is not None else ''

        if row_id:
            current_id = row_id
            current_name = row_name
        elif not row_id and current_id:
            row_id = current_id
            row_name = current_name

        date_str = row[date_col].strip() if date_col is not None else ''

        if not row_id or not date_str:
            continue

        # Create one record per period column
        for period_num, col_idx in sorted(period_cols.items()):
            cell_value = row[col_idx].strip() if col_idx < len(row) else ''
            cell_lower = cell_value.lower()

            status, reason = SYNERGY_STATUS_MAP.get(
                cell_lower, ('absent', cell_value))

            flat_rows.append([
                row_id,       # student_id
                date_str,     # date
                str(period_num),  # period
                status,       # status (already lowercase)
                '',           # course_name
                reason,       # reason (original Synergy value)
            ])

    return flat_rows
