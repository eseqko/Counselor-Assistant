"""Parse student class schedules from Synergy exports into normalized rows.

Two adapters, one output shape, so the import pipeline never has to care which
file it got:

    parse_schedule_excel(file)  ->  [ParsedRow, ...]   (Synergy U-SCH100 export)
    parse_schedule_pdf(file)    ->  [ParsedRow, ...]   (printed first-day PDF)

Excel is the reliable path — exact fields, whole caseload in one file, and it
survives a report-template change. The PDF path exists because a counselor
often has only the printout a student was handed, and it reads the page by
COLUMN POSITION rather than by line text, because the extracted text order does
not match the visual column order.
"""
import re
from dataclasses import dataclass, field
from datetime import date, timedelta

# Rows that are administrative assignments rather than classes. Synergy puts a
# "Vice Principal" row in period 7 with no room; it must not count toward
# credits or a teacher's load, but is worth keeping on the record.
NON_CLASS_TITLES = ('vice principal', 'principal', 'counselor', 'dean',
                    'administrator', 'assistant principal')

ADVISORY_TITLES = ('advisory', 'homeroom', 'advisement', 'flex time')

# Excel serial dates count from 1899-12-30 (the Lotus-compatible epoch Excel uses).
_EXCEL_EPOCH = date(1899, 12, 30)


@dataclass
class ParsedRow:
    """One course enrollment, identical no matter which file it came from."""
    student_ref: str = ''       # Perm ID when present, else the name string
    student_name: str = ''
    grade_level: int = None
    school_year: str = ''
    term: str = ''              # canonical: Q1|Q2|Q3|Q4|YR
    period: int = None
    course_number: str = ''
    course_title: str = ''
    section_id: str = ''
    teacher_name: str = ''
    room: str = ''
    start_date: object = None
    is_advisory: bool = False
    is_non_class: bool = False
    source: str = ''
    warnings: list = field(default_factory=list)

    @property
    def key(self):
        """Identity of an enrollment within one student-year, for dedupe."""
        return (self.school_year, self.term, self.period, self.course_number)


def normalize_term(raw):
    """Map every spelling of a term onto Q1-Q4 / YR.

    Synergy's Excel export writes 'YR' for a year-long row; the printed PDF
    writes 'FALL & SPRING' for the identical thing. Without this the same
    schedule imported two ways would produce two different records.
    """
    t = (raw or '').strip().upper()
    if not t:
        return ''
    m = re.fullmatch(r'Q\s*([1-4])', t)
    if m:
        return f'Q{m.group(1)}'
    if t in ('YR', 'YEAR', 'FULL YEAR', 'FALL & SPRING', 'FALL AND SPRING',
             'FALL/SPRING', 'ALL YEAR'):
        return 'YR'
    if t.startswith('S') and t[1:].strip() in ('1', '2'):
        return t.replace(' ', '')          # S1 / S2 kept as-is
    return t[:12]


def classify_title(title):
    """(is_advisory, is_non_class) for a course title."""
    t = (title or '').strip().lower()
    if not t:
        return False, False
    if any(k in t for k in ADVISORY_TITLES):
        return True, False
    # Match the administrative titles as whole-string-ish, so a real course
    # named "Intro to Counseling" isn't mistaken for a counselor assignment.
    if any(t == k or t.startswith(k) for k in NON_CLASS_TITLES):
        return False, True
    return False, False


def excel_serial_to_date(value):
    """Convert an Excel serial date to a date, tolerating junk."""
    try:
        n = int(float(value))
    except (TypeError, ValueError):
        return None
    if n <= 0 or n > 100000:
        return None
    return _EXCEL_EPOCH + timedelta(days=n)


def _clean(v):
    """Cell value -> trimmed string, without turning 0001 into 1 or 5 into 5.0."""
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


# ── Excel adapter (Synergy U-SCH100) ────────────────────────────────

# Canonical field -> accepted header spellings (lowercased, compared loosely).
EXCEL_COL_ALIASES = {
    'student_ref':   ('perm id', 'permid', 'student id', 'student id #', 'sis id'),
    'student_name':  ('student name', 'studentname', 'name'),
    'grade_level':   ('grade', 'grade level', 'gradelevel', 'grd'),
    'course_number': ('course id', 'courseid', 'course number', 'course #'),
    'course_title':  ('course title', 'coursetitle', 'course name', 'course'),
    'section_id':    ('section id', 'sectionid', 'section'),
    'term':          ('term code', 'termcode', 'term', 'mark name'),
    'period':        ('begin period', 'period', 'per', 'beginperiod'),
    'teacher_name':  ('staff name', 'staffname', 'teacher', 'teacher name'),
    'room':          ('room name', 'roomname', 'room'),
    'start_date':    ('enter date', 'enterdate', 'start date'),
    'school_year':   ('fullyear', 'full year', 'school year', 'schoolyear', 'year'),
}


def build_col_map(header):
    """Map canonical field names to column indices from an actual header row."""
    norm = [re.sub(r'\s+', ' ', str(h or '')).strip().lower() for h in header]
    col_map = {}
    for canon, aliases in EXCEL_COL_ALIASES.items():
        for i, h in enumerate(norm):
            if h in aliases:
                col_map[canon] = i
                break
    return col_map


def rows_to_parsed(header, data_rows, source='excel'):
    """Shared Excel/CSV row -> ParsedRow conversion, given a header row."""
    col = build_col_map(header)
    missing = [f for f in ('course_number', 'period', 'term') if f not in col]
    if missing:
        raise ValueError(
            'This file is missing required column(s): '
            + ', '.join(missing.copy())
            + '. Expected a Synergy schedule export (U-SCH100) with '
              'Course ID, Begin Period and Term Code.')

    def cell(row, name):
        i = col.get(name)
        return _clean(row[i]) if i is not None and i < len(row) else ''

    out = []
    for row in data_rows:
        course_number = cell(row, 'course_number')
        title = cell(row, 'course_title')
        if not course_number and not title:
            continue                       # blank spacer row

        advisory, non_class = classify_title(title)
        period_raw = cell(row, 'period')
        grade_raw = cell(row, 'grade_level')

        pr = ParsedRow(
            student_ref=cell(row, 'student_ref'),
            student_name=cell(row, 'student_name'),
            grade_level=int(grade_raw) if grade_raw.isdigit() else None,
            school_year=cell(row, 'school_year'),
            term=normalize_term(cell(row, 'term')),
            period=int(period_raw) if period_raw.isdigit() else None,
            course_number=course_number,
            course_title=title,
            section_id=cell(row, 'section_id'),
            teacher_name=cell(row, 'teacher_name'),
            room=cell(row, 'room'),
            start_date=excel_serial_to_date(cell(row, 'start_date')),
            is_advisory=advisory,
            is_non_class=non_class,
            source=source,
        )
        if pr.period is None:
            pr.warnings.append('No period on this row')
        out.append(pr)
    return out


def parse_schedule_excel(file):
    """Parse a Synergy schedule export (.xls, .xlsx or .csv)."""
    header, data_rows = read_tabular(file)
    return rows_to_parsed(header, data_rows, source='excel')


def read_tabular(file):
    """Read .xls / .xlsx / .csv into (header, rows).

    Synergy exports legacy BIFF .xls, which openpyxl cannot open at all — it
    raises InvalidFileException and tells you to use xlrd. So the two Excel
    formats need different readers, chosen by sniffing the actual bytes rather
    than trusting the extension (the export is often named .XLS regardless).
    """
    import csv as _csv
    import io as _io

    raw = file.read()
    if hasattr(file, 'seek'):
        file.seek(0)

    # OLE2 compound-document magic => legacy BIFF .xls
    if raw[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
        try:
            import xlrd
        except ImportError:
            raise ValueError(
                'This is a legacy .xls file and the xlrd library is not '
                'installed. Either run "pip install xlrd", or open the file in '
                'Excel and use Save As -> .xlsx.')
        book = xlrd.open_workbook(file_contents=raw)
        sheet = book.sheet_by_index(0)
        grid = [[sheet.cell_value(r, c) for c in range(sheet.ncols)]
                for r in range(sheet.nrows)]
        return (grid[0], grid[1:]) if grid else ([], [])

    # Zip magic => xlsx
    if raw[:2] == b'PK':
        from openpyxl import load_workbook
        wb = load_workbook(_io.BytesIO(raw), data_only=True, read_only=True)
        ws = wb.active
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
        return (grid[0], grid[1:]) if grid else ([], [])

    # Otherwise treat as delimited text.
    text = raw.decode('utf-8-sig', errors='replace')
    grid = list(_csv.reader(_io.StringIO(text)))
    return (grid[0], grid[1:]) if grid else ([], [])


# ── PDF adapter (printed first-day schedule) ────────────────────────

# Column X-position bands measured off the real Synergy schedule printout.
# The extracted text order does NOT match the visual column order, so bucketing
# runs by X is the only reliable way to read this layout.
PDF_COLUMNS = [
    ('period', 60, 108),
    ('course_number', 108, 172),
    ('course_title', 172, 380),
    ('teacher_name', 380, 470),
    ('room', 470, 532),
    ('term', 532, 600),
]
# Rows sit ~15pt apart; runs within this many points share a row.
_ROW_TOLERANCE = 4.0
# Everything below this Y on the page is the footer/legend block.
_FOOTER_Y = 170.0


def parse_schedule_pdf(file):
    """Parse one printed schedule PDF (all pages; one student per page)."""
    from PyPDF2 import PdfReader

    reader = PdfReader(file)
    out = []
    for page in reader.pages:
        out.extend(_parse_pdf_page(page))
    return out


def _parse_pdf_page(page):
    runs = []

    def visitor(text, cm, tm, font_dict, font_size):
        t = (text or '').strip()
        if t:
            runs.append((round(tm[5], 1), round(tm[4], 1), t))

    page.extract_text(visitor_text=visitor)
    if not runs:
        return []

    # Group runs into visual rows by Y, tolerating sub-point jitter.
    rows = []
    for y, x, t in sorted(runs, key=lambda r: (-r[0], r[1])):
        if rows and abs(rows[-1][0] - y) <= _ROW_TOLERANCE:
            rows[-1][1].append((x, t))
        else:
            rows.append([y, [(x, t)]])

    meta = _pdf_page_meta(rows)

    parsed = []
    for y, parts in rows:
        if y < _FOOTER_Y:
            continue                        # footer/legend block
        cells = {name: '' for name, _, _ in PDF_COLUMNS}
        for x, t in sorted(parts):
            for name, lo, hi in PDF_COLUMNS:
                if lo <= x < hi:
                    cells[name] = (cells[name] + ' ' + t).strip()
                    break

        period, course_number = cells['period'], cells['course_number']
        is_row = (re.fullmatch(r'\d{1,2}', period)
                  and re.fullmatch(r'[0-9A-Za-z]{3,10}', course_number))

        if is_row:
            advisory, non_class = classify_title(cells['course_title'])
            parsed.append(ParsedRow(
                student_ref=meta.get('student_ref', ''),
                student_name=meta.get('student_name', ''),
                grade_level=meta.get('grade_level'),
                school_year=meta.get('school_year', ''),
                term=normalize_term(cells['term']),
                period=int(period),
                course_number=course_number,
                course_title=cells['course_title'],
                teacher_name=cells['teacher_name'],
                room=cells['room'],
                is_advisory=advisory,
                is_non_class=non_class,
                source='pdf',
            ))
        elif parsed and cells['course_title'] and not period and not course_number:
            # A long course title wraps onto its own baseline, carrying only
            # the trailing "[S1]" marker. It belongs to the row just above.
            parsed[-1].course_title = (
                parsed[-1].course_title + ' ' + cells['course_title']).strip()

    # Re-classify after any wrapped titles were reassembled.
    for pr in parsed:
        pr.course_title = re.sub(r'\s+', ' ', pr.course_title).strip()
        pr.is_advisory, pr.is_non_class = classify_title(pr.course_title)
    return parsed


# The header has two side-by-side blocks: "Student Information" on the left and
# "Login Information" on the right. They share text baselines, so reading a whole
# line would splice the StudentVUE user id onto the student's name.
_STUDENT_BLOCK_MAX_X = 380.0
# Placeholder fragments Synergy prints in the login block, plus what's left when
# a field has been redacted — never a real name.
_NOT_A_NAME = {'stu.', 'stu', 'n/a', '-', 'confidential'}


def _pdf_page_meta(rows):
    """Pull student/year/grade off the header block of a schedule page.

    Restricted to the left-hand Student Information column: the right-hand
    Login Information block sits on the same baselines, and splicing the two
    produced a "student name" of "stu." (the redacted StudentVUE user id),
    which then silently failed to match instead of prompting for the student.
    """
    meta = {}
    for y, parts in rows:
        left = ' '.join(t for x, t in sorted(parts) if x < _STUDENT_BLOCK_MAX_X)
        full = ' '.join(t for _, t in sorted(parts))

        m = re.search(r'(20\d{2}-20\d{2})', full)
        if m and 'school_year' not in meta:
            meta['school_year'] = m.group(1)
        if 'Current Grade' in left:
            g = re.search(r'Current Grade:?\s*(\d{1,2})', left) or \
                re.search(r'\b(\d{1,2})\b', left)
            if g:
                meta['grade_level'] = int(g.group(1))
        if 'Student ID' in left:
            sid = re.search(r'Student ID:?\s*([A-Za-z0-9\-]+)', left)
            if sid:
                meta['student_ref'] = sid.group(1)
        for label, key in (('Last Name:', 'last'), ('First Name:', 'first')):
            if label in left:
                val = left.split(label, 1)[1].strip()
                val = re.split(r'\s{2,}|[A-Z][a-z]+ Name:', val)[0].strip()
                if val and val.lower() not in _NOT_A_NAME:
                    meta[key] = val

    if meta.get('first') or meta.get('last'):
        meta['student_name'] = f"{meta.get('last', '')}, {meta.get('first', '')}".strip(', ')
    return meta


def parse_schedule_file(file, filename=''):
    """Dispatch on file type and return normalized rows from either adapter."""
    name = (filename or getattr(file, 'filename', '') or '').lower()
    if name.endswith('.pdf'):
        return parse_schedule_pdf(file)
    return parse_schedule_excel(file)
